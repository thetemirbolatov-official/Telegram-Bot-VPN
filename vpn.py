import os
import logging
import time
import json
from datetime import datetime, timedelta
import threading
from telebot import TeleBot, types, util
import requests
import urllib.parse
import qrcode
from io import BytesIO
import webbrowser
import random
import zipfile
import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Конфигурация бота
TOKEN = 'your api token'
ADMIN_ID = #you id 
bot = TeleBot(TOKEN, threaded=True, num_threads=4)
CHANNEL_USERNAME = ""  # Оставить пустым это мусор
CHANNELS = [
    "@you name bot_09",  #добавь свои каналы
    # "@your_channel_2",
]

SUBSCRIPTION_WARNING_DAYS = 3  # За сколько дней предупреждать об окончании

# Промокоды
try:
    with open('promo_codes.json', 'r') as f:
        PROMO_CODES = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    PROMO_CODES = {}

# Настройка логгирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('vpn_bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Загрузка данных
try:
    with open('users_db.json', 'r') as f:
        users_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    users_db = {}

try:
    with open('payments_db.json', 'r') as f:
        payments_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    payments_db = {}

try:
    with open('servers_db.json', 'r') as f:
        servers_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
# Добавляем реальные серверы в базу
    servers_db = {
    'server1': {
        'name': '🇩🇪 Germany', 
        'location': 'Germany, Frankfurt', 
        'load': 'Low', 
        'protocol': 'WireGuard',
        'ip': 'vpn-de1.example.com',
        'available_configs': [],
        'used_configs': {}
    },
    'server2': {
        'name': '🇳🇱 Netherlands', 
        'location': 'Netherlands, Amsterdam', 
        'load': 'Low', 
        'protocol': 'WireGuard',
        'ip': 'vpn-nl1.example.com',
        'available_configs': [],
        'used_configs': {}
    }
}

# Функция для загрузки конфигов сервера
def load_server_configs(server_name, config_files):
    server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
    servers_db[server_key]['available_configs'] = config_files
    save_data_to_file()

# Функция для получения случайного конфига
def get_random_config(server_name, user_id):
    try:
        # Находим сервер по имени
        server_key = None
        for key, server_data in servers_db.items():
            if server_data['name'] == server_name:
                server_key = key
                break
        
        if not server_key:
            logger.error(f"Сервер {server_name} не найден в базе")
            return None
        
        # Проверяем есть ли доступные конфиги
        if not servers_db[server_key]['available_configs']:
            logger.error(f"Нет доступных конфигов для сервера {server_name}")
            logger.error(f"Доступные конфиги: {servers_db[server_key]['available_configs']}")
            return None
        
        # Выбираем случайный конфиг
        config_file = random.choice(servers_db[server_key]['available_configs'])
        
        # Переносим конфиг из доступных в используемые
        servers_db[server_key]['available_configs'].remove(config_file)
        servers_db[server_key]['used_configs'][str(user_id)] = config_file
        save_data_to_file()
        
        logger.info(f"Выдан конфиг {config_file} для пользователя {user_id} на сервере {server_name}")
        return config_file
        
    except Exception as e:
        logger.error(f"Ошибка в get_random_config: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None
    
try:
    with open('payment_methods.json', 'r') as f:
        payment_methods = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    payment_methods = {}
    
# Цены и периоды подписки
SUBSCRIPTION_PLANS = {
    '1 месяц': {'price': '199₽', 'days': 30},
    '3 месяца': {'price': '499₽', 'days': 90},
    '6 месяцев': {'price': '999₽', 'days': 180}
}

# Функции для работы с данными
# Модифицируем функцию save_data_to_file для использования UTF-8 encoding
def save_data_to_file():
    try:
        # Создаем backup перед сохранением
        backup_files = ['users_db.json', 'payments_db.json', 'servers_db.json', 'payment_methods.json', 'promo_codes.json']
        for file in backup_files:
            if os.path.exists(file):
                shutil.copy2(file, f"{file}.backup")
        
        with open('users_db.json', 'w', encoding='utf-8') as f:
            json.dump(users_db, f, ensure_ascii=False, indent=2)
        with open('payments_db.json', 'w', encoding='utf-8') as f:
            json.dump(payments_db, f, ensure_ascii=False, indent=2)
        with open('servers_db.json', 'w', encoding='utf-8') as f:
            json.dump(servers_db, f, ensure_ascii=False, indent=2)
        with open('payment_methods.json', 'w', encoding='utf-8') as f:
            json.dump(payment_methods, f, ensure_ascii=False, indent=2)
        with open('promo_codes.json', 'w', encoding='utf-8') as f:
            json.dump(PROMO_CODES, f, ensure_ascii=False, indent=2)
        
        logger.info("Данные успешно сохранены")
    except Exception as e:
        logger.error(f"Ошибка сохранения данных: {e}")
        # Восстанавливаем из backup при ошибке
        try:
            for file in backup_files:
                backup_file = f"{file}.backup"
                if os.path.exists(backup_file):
                    shutil.copy2(backup_file, file)
            logger.info("Данные восстановлены из backup")
        except Exception as backup_error:
            logger.error(f"Ошибка восстановления из backup: {backup_error}")

# Модифицируем загрузку данных в начале файла для использования UTF-8
try:
    with open('users_db.json', 'r', encoding='utf-8') as f:
        users_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    users_db = {}

try:
    with open('payments_db.json', 'r', encoding='utf-8') as f:
        payments_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    payments_db = {}

try:
    with open('servers_db.json', 'r', encoding='utf-8') as f:
        servers_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    # Добавляем реальные серверы в базу
    servers_db = {
        'server1': {
            'name': '🇩🇪 Germany', 
            'location': 'Germany, Frankfurt', 
            'load': 'Low', 
            'protocol': 'WireGuard',
            'ip': 'vpn-de1.example.com',
            'available_configs': [],
            'used_configs': {}
        },
        'server2': {
            'name': '🇳🇱 Netherlands', 
            'location': 'Netherlands, Amsterdam', 
            'load': 'Low', 
            'protocol': 'WireGuard',
            'ip': 'vpn-nl1.example.com',
            'available_configs': [],
            'used_configs': {}
        }
    }

try:
    with open('payment_methods.json', 'r', encoding='utf-8') as f:
        payment_methods = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    payment_methods = {}

try:
    with open('promo_codes.json', 'r', encoding='utf-8') as f:
        PROMO_CODES = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    PROMO_CODES = {}

def save_user_data(user_id, data):
    users_db[str(user_id)] = data
    save_data_to_file()

def get_user_data(user_id):
    return users_db.get(str(user_id), {})

def save_payment(payment_id, data):
    payments_db[payment_id] = data
    save_data_to_file()

def get_payment(payment_id):
    return payments_db.get(payment_id)

def generate_payment_id():
    return str(random.randint(100, 999))  # 3 цифры

def is_admin(user_id):
    return str(user_id) == str(ADMIN_ID)

def subscription_monitor():
    """Фоновая задача для мониторинга подписок"""
    while True:
        try:
            # Проверяем подписки каждые 6 часов
            time.sleep(6 * 60 * 60)  # 6 часов в секундах
            
            # Проверяем уведомления об окончании
            check_and_send_subscription_warnings()
            
            # Проверяем истекшие подписки
            expired = check_expired_subscriptions()
            if expired:
                logger.info(f"Найдено {len(expired)} просроченных подписок")
                
        except Exception as e:
            logger.error(f"Ошибка в subscription_monitor: {e}")
            time.sleep(300)  # Ждем 5 минут при ошибке


def create_config_file(server_name, user_id=None, config_content=None):
    config_filename = f"{server_name}_{user_id}.conf" if user_id else f"{server_name}.conf"
    try:
        with open(config_filename, 'w') as f:
            if config_content:
                f.write(config_content)
            else:
                private_key = f"user_{user_id}_private_key" if user_id else "your_private_key"
                public_key = f"server_{server_name}_public_key"
                f.write(f"[Interface]\nPrivateKey = {private_key}\nAddress = 10.0.0.1/24\n\n[Peer]\nPublicKey = {public_key}\nAllowedIPs = 0.0.0.0/0\nEndpoint = {server_name}:51820")
        return config_filename
    except Exception as e:
        logger.error(f"Ошибка создания конфига: {e}")
        return None

def delete_previous_message(chat_id, message_id=None):
    try:
        if message_id and isinstance(message_id, int):
            bot.delete_message(chat_id, message_id)
    except Exception as e:
        error_msg = str(e).lower()
        # Игнорируем распространенные ошибки удаления
        if any(phrase in error_msg for phrase in [
            "message to delete not found",
            "message can't be deleted",
            "bad request: message can't be deleted"
        ]):
            logger.debug(f"Не удалось удалить сообщение (нормально): {e}")
        else:
            logger.error(f"Ошибка удаления сообщения: {e}")

def check_expired_subscriptions():
    """Проверяет истекшие подписки и отправляет уведомления пользователям и админу"""
    expired_users = []
    current_time = datetime.now()
    
    for user_id, user_data in users_db.items():
        if 'subscriptions' in user_data:
            for sub in user_data['subscriptions']:
                expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                
                # Проверяем истекла ли подписка
                if expiry_date < current_time:
                    # Проверяем не отправляли ли уже уведомление об окончании
                    if not sub.get('expiry_notification_sent', False):
                        expired_user_info = {
                            'user_id': user_id,
                            'username': user_data.get('username', 'N/A'),
                            'config_file': sub.get('config_file', 'N/A'),
                            'server': sub.get('server', 'N/A'),
                            'expiry_date': sub['expiry_date']
                        }
                        expired_users.append(expired_user_info)
                        
                        # Отправляем уведомление пользователю
                        send_subscription_expired_notification(user_id, sub)
                        
                        # Отправляем уведомление админу
                        admin_notification = f"""⚠️ ПРОСРОЧЕНА ПОДПИСКА ⚠️

👤 Пользователь: @{expired_user_info['username']} (ID: {user_id})
🖥 Сервер: {expired_user_info['server']}
🔑 Конфиг: {expired_user_info['config_file']}
📅 Истекла: {expired_user_info['expiry_date']}

Для проверки перейдите в раздел:
🔍 Просроченные подписки"""

                        try:
                            bot.send_message(ADMIN_ID, admin_notification)
                        except Exception as e:
                            logger.error(f"Ошибка отправки уведомления админу: {e}")
                        
                        # Помечаем что уведомление отправлено
                        sub['expiry_notification_sent'] = True
    
    # Сохраняем изменения
    if expired_users:
        save_data_to_file()
        logger.info(f"Найдено {len(expired_users)} просроченных подписок. Уведомления отправлены.")
    
    return expired_users

# Клавиатуры
# Клавиатуры
def main_menu_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('🛒 Купить VPN | Продлить')
    )
    keyboard.row(
        types.KeyboardButton('Активировать промокод'),
        types.KeyboardButton('🔑 Мои ключи')
    )
    keyboard.row(
        types.KeyboardButton('📱 Инструкция Android'),  # Новая кнопка
        types.KeyboardButton('📱 Инструкция iPhone')    # Новая кнопка
    )
    keyboard.row(
        types.KeyboardButton('🛟 Поддержка'),
        types.KeyboardButton('⭐⭐⭐ Отзывы')
    )
    return keyboard

# Список каналов для проверки подписки (бот должен быть админом в этих каналах)


# Функция проверки подписки
def check_subscription(user_id):
    """Проверяет подписку пользователя на все каналы из CHANNELS"""
    for channel in CHANNELS:
        try:
            chat_member = bot.get_chat_member(chat_id=channel, user_id=user_id)
            if chat_member.status not in ['member', 'administrator', 'creator']:
                return False
        except Exception as e:
            logger.error(f"Ошибка проверки канала {channel}: {e}")
            return False
    return True

def subscription_keyboard():
    """Клавиатура для подписки на каналы"""
    markup = types.InlineKeyboardMarkup()
    for channel in CHANNELS:
        if isinstance(channel, str) and channel.startswith("@"):
            url = f"https://t.me/{channel.replace('@', '')}"
            btn = types.InlineKeyboardButton(text=f"Подписаться", url=url)
            markup.add(btn)
    btn_check = types.InlineKeyboardButton(text="✅ Я подписался", callback_data="check_sub")
    markup.add(btn_check)
    return markup

# Добавляем кнопку в админ-меню
def admin_menu_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('📊 Статистика'),
        types.KeyboardButton('🔁 Статистика продлений'),
        types.KeyboardButton('📝 Список серверов'),
        types.KeyboardButton('🧾 Проверить платежи'),
        types.KeyboardButton('📢 Рассылка'),
        types.KeyboardButton('⚙️ Настройки оплаты'),
        types.KeyboardButton('🗂 Управление конфигами'),
        types.KeyboardButton('👥 Список покупателей'),
        types.KeyboardButton('🎁 Управление промокодами'),
        types.KeyboardButton('💾 Создать резервную копию'),
        types.KeyboardButton('📥 Восстановить из копии'),
        types.KeyboardButton('📊 Экспорт в Excel'),
        types.KeyboardButton('🗑 Удалить пользователя'),
        types.KeyboardButton('🔄 Перезагрузить конфиги'),
        types.KeyboardButton('🔄 Синхронизировать конфиги'),
        types.KeyboardButton('👤 Список пользователей'),  # Новая кнопка
        types.KeyboardButton('📤 Добавить конфиги массово')  # Новая кнопка
    )
    return keyboard
    
# Добавляем функцию экспорта в Excel
def export_to_excel():
    """Экспортирует все данные в Excel файл"""
    try:
        # Создаем папку для экспорта если ее нет
        if not os.path.exists('exports'):
            os.makedirs('exports')
        
        # Создаем имя файла с timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"exports/data_export_{timestamp}.xlsx"
        
        # Создаем Excel workbook
        wb = Workbook()
        
        # Удаляем дефолтный лист
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 1. Лист с пользователями
        if users_db:
            ws_users = wb.create_sheet("Пользователи")
            users_data = []
            
            for user_id, user_data in users_db.items():
                user_row = {
                    'ID пользователя': user_id,
                    'Username': user_data.get('username', ''),
                    'Подписок': len(user_data.get('subscriptions', [])),
                    'Использованные промокоды': ', '.join(user_data.get('used_promo_codes', []))
                }
                users_data.append(user_row)
            
            df_users = pd.DataFrame(users_data)
            for row in dataframe_to_rows(df_users, index=False, header=True):
                ws_users.append(row)
            
            # Форматируем заголовки
            for cell in ws_users[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Авто-ширина колонок
            for column in ws_users.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_users.column_dimensions[column_letter].width = adjusted_width
        
        # 2. Лист с подписками
        if users_db:
            ws_subs = wb.create_sheet("Подписки")
            subscriptions_data = []
            
            for user_id, user_data in users_db.items():
                for sub in user_data.get('subscriptions', []):
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    is_active = expiry_date > datetime.now()
                    
                    sub_row = {
                        'ID пользователя': user_id,
                        'Username': user_data.get('username', ''),
                        'Сервер': sub.get('server', ''),
                        'Конфиг файл': sub.get('config_file', ''),
                        'Дата покупки': sub.get('purchase_date', ''),
                        'Действует до': sub.get('expiry_date', ''),
                        'Статус': 'Активна' if is_active else 'Истекла',
                        'Тип': sub.get('type', 'платная'),
                        'Промокод': sub.get('promo_code', '')
                    }
                    subscriptions_data.append(sub_row)
            
            df_subs = pd.DataFrame(subscriptions_data)
            for row in dataframe_to_rows(df_subs, index=False, header=True):
                ws_subs.append(row)
            
            # Форматируем заголовки
            for cell in ws_subs[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Авто-ширина колонок
            for column in ws_subs.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_subs.column_dimensions[column_letter].width = adjusted_width
        
        # 3. Лист с платежами
        if payments_db:
            ws_payments = wb.create_sheet("Платежи")
            payments_data = []
            
            for payment_id, payment_data in payments_db.items():
                payment_row = {
                    'ID платежа': payment_id,
                    'ID пользователя': payment_data.get('user_id', ''),
                    'Username': payment_data.get('username', ''),
                    'Сервер': payment_data.get('server', ''),
                    'Срок': payment_data.get('duration', ''),
                    'Сумма': payment_data.get('amount', ''),
                    'Банк': payment_data.get('bank', ''),
                    'Статус': payment_data.get('status', ''),
                    'Дата создания': payment_data.get('timestamp', ''),
                    'Подтвержден': payment_data.get('approved_at', ''),
                    'Кем подтвержден': payment_data.get('approved_by', '')
                }
                payments_data.append(payment_row)
            
            df_payments = pd.DataFrame(payments_data)
            for row in dataframe_to_rows(df_payments, index=False, header=True):
                ws_payments.append(row)
            
            # Форматируем заголовки
            for cell in ws_payments[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Авто-ширина колонок
            for column in ws_payments.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_payments.column_dimensions[column_letter].width = adjusted_width
        
        # 4. Лист с серверами
        if servers_db:
            ws_servers = wb.create_sheet("Серверы")
            servers_data = []
            
            for server_id, server_data in servers_db.items():
                server_row = {
                    'ID сервера': server_id,
                    'Название': server_data.get('name', ''),
                    'Локация': server_data.get('location', ''),
                    'Нагрузка': server_data.get('load', ''),
                    'Протокол': server_data.get('protocol', ''),
                    'IP адрес': server_data.get('ip', ''),
                    'Доступно конфигов': len(server_data.get('available_configs', [])),
                    'Используется конфигов': len(server_data.get('used_configs', {}))
                }
                servers_data.append(server_row)
            
            df_servers = pd.DataFrame(servers_data)
            for row in dataframe_to_rows(df_servers, index=False, header=True):
                ws_servers.append(row)
            
            # Форматируем заголовки
            for cell in ws_servers[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Авто-ширина колонок
            for column in ws_servers.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_servers.column_dimensions[column_letter].width = adjusted_width
        
        # 5. Лист с промокодами
        if PROMO_CODES:
            ws_promo = wb.create_sheet("Промокоды")
            promo_data = []
            
            for code, promo_info in PROMO_CODES.items():
                promo_row = {
                    'Промокод': code,
                    'Сервер': promo_info.get('server', ''),
                    'Дней': promo_info.get('days', ''),
                    'Создан': promo_info.get('created_at', ''),
                    'Создатель': promo_info.get('created_by', '')
                }
                promo_data.append(promo_row)
            
            df_promo = pd.DataFrame(promo_data)
            for row in dataframe_to_rows(df_promo, index=False, header=True):
                ws_promo.append(row)
            
            # Форматируем заголовки
            for cell in ws_promo[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Авто-ширина колонок
            for column in ws_promo.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_promo.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(excel_filename)
        return excel_filename
        
    except Exception as e:
        logger.error(f"Ошибка экспорта в Excel: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None
    
# Добавляем функцию создания резервной копии
def create_backup():
    """Создает резервную копию всех данных"""
    try:
        # Создаем папку для резервных копий если ее нет
        if not os.path.exists('backups'):
            os.makedirs('backups')
        
        # Создаем имя файла с timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backups/backup_{timestamp}.zip"
        
        # Файлы для резервного копирования
        files_to_backup = [
            'users_db.json', 
            'payments_db.json', 
            'servers_db.json',
            'payment_methods.json',
            'promo_codes.json'
        ]
        
        # Добавляем конфигурационные файлы
        config_files = [f for f in os.listdir() if f.endswith('.conf')]
        files_to_backup.extend(config_files)
        
        # Создаем zip архив
        with zipfile.ZipFile(backup_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file in files_to_backup:
                if os.path.exists(file):
                    zipf.write(file)
        
        return backup_filename
    except Exception as e:
        logger.error(f"Ошибка создания резервной копии: {e}")
        return None
        
# Добавляем функцию восстановления из резервной копии
def restore_from_backup(backup_file):
    """Восстанавливает данные из резервной копии с merge"""
    try:
        # Временная папка для распаковки
        temp_dir = 'temp_restore'
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.makedirs(temp_dir)
        
        # Распаковываем архив
        with zipfile.ZipFile(backup_file, 'r') as zipf:
            zipf.extractall(temp_dir)
        
        # Функция для безопасного merge JSON данных
        def merge_json_data(current_data, new_data, key_field=None):
            """Безопасное слияние JSON данных"""
            if isinstance(current_data, dict) and isinstance(new_data, dict):
                # Для словарей объединяем ключи
                result = current_data.copy()
                for key, value in new_data.items():
                    if key in result:
                        if isinstance(result[key], (dict, list)) and isinstance(value, (dict, list)):
                            result[key] = merge_json_data(result[key], value, key_field)
                        else:
                            result[key] = value
                    else:
                        result[key] = value
                return result
            elif isinstance(current_data, list) and isinstance(new_data, list) and key_field:
                # Для списков объединяем по ключу
                result = current_data.copy()
                existing_keys = {item[key_field]: item for item in result if key_field in item}
                
                for new_item in new_data:
                    if key_field in new_item and new_item[key_field] in existing_keys:
                        # Обновляем существующий элемент
                        index = next(i for i, item in enumerate(result) 
                                   if item.get(key_field) == new_item[key_field])
                        result[index] = merge_json_data(result[index], new_item, key_field)
                    else:
                        # Добавляем новый элемент
                        result.append(new_item)
                return result
            else:
                # Для простых типов или несовместимых структур возвращаем новые данные
                return new_data
        
        # Восстанавливаем JSON файлы с merge
        json_files = {
            'users_db.json': 'user_id',
            'payments_db.json': None,  # Простая замена
            'servers_db.json': 'name',
            'payment_methods.json': 'bank',
            'promo_codes.json': None   # Простая замена
        }
        
        for json_file, merge_key in json_files.items():
            temp_file = os.path.join(temp_dir, json_file)
            if os.path.exists(temp_file):
                with open(temp_file, 'r', encoding='utf-8') as f:
                    new_data = json.load(f)
                
                current_file = json_file
                if os.path.exists(current_file):
                    with open(current_file, 'r', encoding='utf-8') as f:
                        current_data = json.load(f)
                    
                    if merge_key:
                        merged_data = merge_json_data(current_data, new_data, merge_key)
                    else:
                        merged_data = new_data  # Простая замена
                else:
                    merged_data = new_data
                
                # Сохраняем объединенные данные
                with open(current_file, 'w', encoding='utf-8') as f:
                    json.dump(merged_data, f, ensure_ascii=False, indent=2)
        
        # Восстанавливаем конфигурационные файлы
        for file in os.listdir(temp_dir):
            if file.endswith('.conf') and not os.path.exists(file):
                shutil.move(os.path.join(temp_dir, file), file)
        
        # Очищаем временную папку
        shutil.rmtree(temp_dir)
        
        # Перезагружаем данные в память
        reload_all_data()
        
        return True
    except Exception as e:
        logger.error(f"Ошибка восстановления из резервной копии: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return False
        
# Добавляем функцию перезагрузки всех данных
def reload_all_data():
    """Перезагружает все данные из файлов"""
    global users_db, payments_db, servers_db, payment_methods, PROMO_CODES
    
    try:
        # Загрузка данных пользователей
        with open('users_db.json', 'r', encoding='utf-8') as f:
            users_db = json.load(f)
        
        # Загрузка данных платежей
        with open('payments_db.json', 'r', encoding='utf-8') as f:
            payments_db = json.load(f)
        
        # Загрузка данных серверов
        with open('servers_db.json', 'r', encoding='utf-8') as f:
            servers_db = json.load(f)
        
        # Загрузка методов оплаты
        with open('payment_methods.json', 'r', encoding='utf-8') as f:
            payment_methods = json.load(f)
        
        # Загрузка промокодов
        with open('promo_codes.json', 'r', encoding='utf-8') as f:
            PROMO_CODES = json.load(f)
            
        logger.info("Все данные успешно перезагружены")
    except Exception as e:
        logger.error(f"Ошибка перезагрузки данных: {e}")

def servers_menu_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    keyboard.add(types.KeyboardButton('🇩🇪 Germany'))
    keyboard.add(types.KeyboardButton('🇳🇱 Netherlands'))
    keyboard.add(types.KeyboardButton('🔙 Назад'))
    return keyboard

def payment_methods_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    
    # Группируем методы по банкам
    bank_groups = {}
    for method in payment_methods.values():
        if method['bank'] not in bank_groups:
            bank_groups[method['bank']] = []
        bank_groups[method['bank']].append(method)
    
    # Создаем кнопки для каждого банка (без указания карт)
    buttons = [types.KeyboardButton(bank) for bank in bank_groups.keys()]
    keyboard.add(*buttons)
    keyboard.add(types.KeyboardButton('🔙 Назад'))
    
    return keyboard

def payment_verification_keyboard(payment_id):
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton(f'✅ Подтвердить {payment_id}'),
        types.KeyboardButton(f'❌ Отклонить {payment_id}'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard

def duration_menu_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('1 месяц - 199₽'),
        types.KeyboardButton('3 месяца - 499₽'),
        types.KeyboardButton('6 месяцев - 999₽'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard

def config_actions_keyboard(config_path):
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('📲 Установить приложение'),
        types.KeyboardButton(f'⚙️ Импортировать {os.path.basename(config_path)}'),
        types.KeyboardButton('💾 Скачать конфиг'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard

def config_management_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('📤 Загрузить новый конфиг'),
        types.KeyboardButton('🗑 Удалить конфиг'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard

def customers_list_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('🔍 Просроченные подписки'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard
    
def check_channel_subscription(user_id):
    """Совместимость со старым кодом - перенаправляет на новую проверку"""
    return check_subscription(user_id)

def check_and_send_subscription_warnings():
    """Проверяет подписки и отправляет уведомления об скором окончании"""
    try:
        current_time = datetime.now()
        warned_users = set()  # Чтобы не дублировать уведомления
        
        for user_id, user_data in users_db.items():
            if 'subscriptions' not in user_data:
                continue
                
            for sub in user_data['subscriptions']:
                expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                days_left = (expiry_date - current_time).days
                
                # Проверяем нужно ли отправлять уведомление
                if 0 < days_left <= SUBSCRIPTION_WARNING_DAYS:
                    # Проверяем не отправляли ли уже уведомление за сегодня
                    warning_key = f"warning_{user_id}_{sub['config_file']}_{days_left}"
                    
                    if warning_key not in sub.get('last_warnings', []):
                        # Отправляем уведомление
                        try:
                            warning_text = f"""⚠️ <b>Внимание!</b>

Ваша подписка на сервере {sub['server']} заканчивается через <b>{days_left}</b> дней.

Дата окончания: {expiry_date.strftime('%d.%m.%Y')}

Чтобы продолжить пользоваться VPN без перерывов, продлите подписку заранее.

🔄 <b>Продлить сейчас:</b> /buy"""
                            
                            bot.send_message(
                                user_id,
                                warning_text,
                                parse_mode='HTML'
                            )
                            
                            # Сохраняем информацию об отправленном уведомлении
                            if 'last_warnings' not in sub:
                                sub['last_warnings'] = []
                            sub['last_warnings'].append(warning_key)
                            warned_users.add(user_id)
                            
                        except Exception as e:
                            logger.error(f"Ошибка отправки уведомления пользователю {user_id}: {e}")
        
        # Сохраняем данные если были отправлены уведомления
        if warned_users:
            save_data_to_file()
            logger.info(f"Отправлены уведомления об окончании подписки для {len(warned_users)} пользователей")
            
    except Exception as e:
        logger.error(f"Ошибка в check_and_send_subscription_warnings: {e}")
        
def send_subscription_expired_notification(user_id, subscription):
    """Отправляет уведомление об окончании подписки"""
    try:
        expiry_date = datetime.strptime(subscription['expiry_date'], "%Y-%m-%d %H:%M:%S")
        
        expired_text = f"""❌ <b>Подписка закончилась</b>

Ваша подписка на сервере {subscription['server']} закончилась {expiry_date.strftime('%d.%m.%Y')}.

Для возобновления работы VPN приобретите новую подписку.

🛒 <b>Приобрести новую подписку:</b> /buy"""

        bot.send_message(
            user_id,
            expired_text,
            parse_mode='HTML'
        )
        
    except Exception as e:
        logger.error(f"Ошибка отправки уведомления об окончании подписки пользователю {user_id}: {e}")
        

        
def generate_qr_code(config_content, filename):
    try:
        # Создаем QR-код
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(config_content)
        qr.make(fit=True)
        
        # Создаем изображение QR-кода
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Сохраняем изображение
        img.save(filename)
        return filename
    except Exception as e:
        logger.error(f"Ошибка генерации QR-кода: {e}")
        return None
        
@bot.message_handler(func=lambda message: message.text == '👤 Список пользователей' and is_admin(message.from_user.id))
def users_list_handler(message):
    """Показывает полный список всех пользователей с детальной информацией"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        if not users_db:
            bot.send_message(message.chat.id, "❌ В базе нет пользователей.")
            return
        
        current_time = datetime.now()
        all_users = []
        
        # Собираем информацию о всех пользователях
        for user_id, user_data in users_db.items():
            user_info = {
                'user_id': user_id,
                'username': user_data.get('username', 'N/A'),
                'subscriptions': []
            }
            
            if 'subscriptions' in user_data:
                for sub in user_data['subscriptions']:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    days_left = (expiry_date - current_time).days
                    is_expired = expiry_date < current_time
                    
                    subscription_info = {
                        'server': sub['server'],
                        'config_file': sub['config_file'],
                        'purchase_date': sub.get('purchase_date', 'N/A'),
                        'expiry_date': sub['expiry_date'],
                        'days_left': days_left,
                        'is_expired': is_expired,
                        'expired_days': (current_time - expiry_date).days if is_expired else 0
                    }
                    user_info['subscriptions'].append(subscription_info)
            
            all_users.append(user_info)
        
        # Сортируем: сначала пользователи с просроченными подписками, затем по оставшимся дням
        def sort_key(user):
            if not user['subscriptions']:
                return (1, 9999)  # Пользователи без подписок в конце
            has_expired = any(sub['is_expired'] for sub in user['subscriptions'])
            if has_expired:
                # Для пользователей с просроченными подписками - по количеству дней просрочки
                expired_subs = [sub for sub in user['subscriptions'] if sub['is_expired']]
                max_expired_days = max(sub['expired_days'] for sub in expired_subs)
                return (0, -max_expired_days)  # Больше дней просрочки - выше в списке
            else:
                # Для активных пользователей - по минимальному количеству оставшихся дней
                min_days_left = min(sub['days_left'] for sub in user['subscriptions'])
                return (1, min_days_left)
        
        all_users.sort(key=sort_key)
        
        # Статистика
        total_users = len(all_users)
        expired_count = sum(1 for user in all_users if any(sub['is_expired'] for sub in user['subscriptions']))
        active_count = total_users - expired_count
        
        # Отправляем общую статистику сначала
        stats_text = f"👥 <b>Полный список пользователей</b>\n\n"
        stats_text += f"📊 Всего пользователей: {total_users}\n"
        stats_text += f"✅ Активных подписок: {active_count}\n"
        stats_text += f"⚠️ Просроченных подписок: {expired_count}\n\n"
        stats_text += "────────────────────\n\n"
        
        bot.send_message(message.chat.id, stats_text, parse_mode='HTML')
        
        # Отправляем пользователей частями (по 3 пользователя в сообщении)
        users_per_message = 3
        total_parts = (len(all_users) - 1) // users_per_message + 1
        
        for part_num in range(total_parts):
            start_idx = part_num * users_per_message
            end_idx = start_idx + users_per_message
            chunk = all_users[start_idx:end_idx]
            
            text = f"📄 <b>Часть {part_num + 1}/{total_parts}</b>\n\n"
            
            for user in chunk:
                text += f"🆔 <b>ID:</b> <code>{user['user_id']}</code>\n"
                text += f"👤 <b>Username:</b> @{user['username']}\n"
                
                if not user['subscriptions']:
                    text += "❌ <b>Нет активных подписок</b>\n"
                else:
                    # Показываем только первую подписку для экономии места
                    sub = user['subscriptions'][0]
                    status_icon = "❌" if sub['is_expired'] else "✅"
                    status_text = f"<b>ПРОСРОЧЕН</b> ({sub['expired_days']} дней)" if sub['is_expired'] else f"активен ({sub['days_left']} дней)"
                    
                    text += f"{status_icon} <b>Сервер:</b> {sub['server']}\n"
                    text += f"🔑 <b>Конфиг:</b> {os.path.basename(sub['config_file'])}\n"
                    text += f"📅 <b>До:</b> {sub['expiry_date']}\n"
                    text += f"📊 <b>Статус:</b> {status_text}\n"
                    
                    # Если подписок больше одной, показываем количество
                    if len(user['subscriptions']) > 1:
                        text += f"📋 <b>Всего подписок:</b> {len(user['subscriptions'])}\n"
                
                text += "────────────────────\n\n"
            
            # Проверяем длину сообщения и разбиваем если нужно
            if len(text) > 4000:
                # Разбиваем на части по строкам
                lines = text.split('\n')
                current_part = ""
                for line in lines:
                    if len(current_part + line + '\n') > 4000:
                        bot.send_message(message.chat.id, current_part, parse_mode='HTML')
                        current_part = line + '\n'
                    else:
                        current_part += line + '\n'
                if current_part:
                    bot.send_message(message.chat.id, current_part, parse_mode='HTML')
            else:
                bot.send_message(message.chat.id, text, parse_mode='HTML')
            
            # Небольшая пауза между сообщениями
            time.sleep(0.3)
        
        # Добавляем финальное сообщение с инструкциями
        instructions_text = (
            "💡 <b>Инструкция для администратора:</b>\n\n"
            "• Для удаления пользователя используйте ID из списка\n"
            "• Просроченные подписки выделены ❌\n" 
            "• Активные подписки выделены ✅\n"
            "• Для детального просмотра используйте кнопку '👥 Список покупателей'\n\n"
            "🛠 <b>Быстрые действия:</b>"
        )
        
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('🗑 Удалить пользователя'))
        keyboard.add(types.KeyboardButton('🔍 Просроченные подписки'))
        keyboard.add(types.KeyboardButton('👥 Список покупателей'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(
            message.chat.id,
            instructions_text,
            parse_mode='HTML',
            reply_markup=keyboard
        )
        
    except Exception as e:
        logger.error(f"Ошибка в users_list_handler: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        bot.send_message(message.chat.id, "❌ Ошибка при загрузке списка пользователей")
        
@bot.message_handler(func=lambda message: message.text == 'Инструкция установки 🎬')
def installation_instructions(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        # Текст инструкции (ваш текст с смайликами)
        instructions_text = """📋 <b>Инструкция установки:</b>

1️⃣ <b>Скачайте приложение AmneziaWG</b> (ссылка ниже 👇)

2️⃣ <b>Внутри бота купите</b> 🛒 <b>или активируйте промокод</b> 🎫

3️⃣ <b>Вставьте ключ</b> 🔑 <b>который прислал бот в приложение AmneziaWG</b>"""
        
        # Отправляем фото с инструкцией
        try:
            with open('Inst.png', 'rb') as photo:
                bot.send_photo(
                    user_id,
                    photo,
                    caption=instructions_text,
                    parse_mode='HTML'
                )
        except FileNotFoundError:
            # Если фото нет, отправляем только текст
            bot.send_message(
                user_id,
                instructions_text,
                parse_mode='HTML'
            )
        
        # Второе сообщение с кнопками для скачивания
        download_text = "📥 <b>Скачать приложение AmneziaWG:</b>"
        
        # Создаем инлайн клавиатуру для скачивания приложения
        markup = types.InlineKeyboardMarkup()
        markup.row(
            types.InlineKeyboardButton("Android", url="https://play.google.com/store/apps/details?id=org.amnezia.awg"),
            types.InlineKeyboardButton("iOS", url="https://apps.apple.com/ru/app/amneziawg/id6478942365")
        )
        markup.row(
            types.InlineKeyboardButton("Windows", url="https://github.com/amnezia-vpn/amneziawg-windows-client/releases/tag/1.0.2"),
            types.InlineKeyboardButton("macOS", url="https://apps.apple.com/us/app/amneziawg/id6478942365")
        )
        
        bot.send_message(
            user_id,
            download_text,
            parse_mode='HTML',
            reply_markup=markup
        )
            
    except Exception as e:
        logger.error(f"Ошибка в installation_instructions: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка при загрузке инструкции.")
        
@bot.message_handler(func=lambda message: message.text == '🔗 ссылка на приложение')
def download_app_handler(message):
    """Обработчик кнопки скачивания приложения после активации промокода"""
    try:
        user_id = message.from_user.id
        
        download_text = "📥 <b>Скачать приложение AmneziaWG:</b>"
        markup = types.InlineKeyboardMarkup()
        markup.row(
            types.InlineKeyboardButton("Android", url="https://play.google.com/store/apps/details?id=org.amnezia.awg"),
            types.InlineKeyboardButton("iOS", url="https://apps.apple.com/ru/app/amneziawg/id6478942365")
        )
        markup.row(
            types.InlineKeyboardButton("Windows", url="https://github.com/amnezia-vpn/amneziawg-windows-client/releases/tag/1.0.2"),
            types.InlineKeyboardButton("macOS", url="https://apps.apple.com/us/app/amneziawg/id6478942365")
        )
        
        bot.send_message(
            user_id,
            download_text,
            parse_mode='HTML',
            reply_markup=markup
        )
        

        
    except Exception as e:
        logger.error(f"Ошибка в download_app_handler: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при отправке ссылок на приложение.")

@bot.message_handler(func=lambda message: message.text == '📱 Инструкция Android')
def android_instruction(message):
    try:
        instruction_text = """📱 Инструкция для Android:

1. Получить файл you name bot.conf у бота
2. Скачать приложение AmneziaWG
3. Импортировать файл you name bot.conf в приложение
4. Дать все необходимые разрешения
5. Ввести пароль от телефона (если требуется)
6. Готово! VPN успешно подключен

📹 Видеоинструкция: https://t.me/you name bot_09/38"""
        
        bot.send_message(message.chat.id, instruction_text)
    except Exception as e:
        logger.error(f"Ошибка в android_instruction: {e}")

@bot.message_handler(func=lambda message: message.text == '📱 Инструкция iPhone')
def iphone_instruction(message):
    try:
        instruction_text = """📱 Инструкция для iPhone:

1. Получить файл you name bot.conf у бота
2. Скачать приложение AmneziaWG из App Store
3. Нажать на файл you name bot.conf
4. Внизу нажать кнопку "Поделиться"
5. Из списка приложений выбрать AmneziaWG
6. Готово! VPN успешно подключен

📹 Видеоинструкция: https://t.me/you name bot_09/41"""
        
        bot.send_message(message.chat.id, instruction_text)
    except Exception as e:
        logger.error(f"Ошибка в iphone_instruction: {e}")        
                        

    
# Добавляем обработчик для кнопки экспорта
@bot.message_handler(func=lambda message: message.text == '📊 Экспорт в Excel' and is_admin(message.from_user.id))
def export_excel_handler(message):
    try:
        bot.send_message(message.chat.id, "📊 Создаю Excel отчет... Это может занять несколько секунд.")
        
        excel_file = export_to_excel()
        if excel_file:
            with open(excel_file, 'rb') as f:
                bot.send_document(
                    message.chat.id,
                    f,
                    caption="✅ Excel отчет успешно создан!\n\n"
                           "📋 Содержит листы:\n"
                           "• Пользователи\n"
                           "• Подписки\n"
                           "• Платежи\n"
                           "• Серверы\n"
                           "• Промокоды",
                    visible_file_name=os.path.basename(excel_file)
                )
        else:
            bot.send_message(message.chat.id, "❌ Ошибка создания Excel отчета")
    except Exception as e:
        logger.error(f"Ошибка в export_excel_handler: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка создания Excel отчета")
        
    # Обработчики команд
@bot.message_handler(commands=['start', 'help'])
def start_handler(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        # Проверяем подписку
        if not check_subscription(user_id):
            bot.send_message(
                user_id,
                "👋 Для использования бота необходимо подписаться на наши каналы:",
                reply_markup=subscription_keyboard()
            )
            return
        
        # Если подписка есть - показываем приветственное сообщение
        if is_admin(user_id):
            bot.send_message(user_id, "👋 Добро пожаловать в админ-панель!", reply_markup=admin_menu_keyboard())
        else:
            try:
                with open('welcome.png', 'rb') as photo:
                    bot.send_photo(
                        user_id, 
                        photo,
                        caption="""💙 Добро пожаловать в you name bot VPN  

Защита и свобода в сети —  
невидимость, доступ, скорость.  
Выбери функцию:
""",
                        reply_markup=main_menu_keyboard()
                    )
            except FileNotFoundError:
                bot.send_message(
                    user_id,
                    """💙 Добро пожаловать в you name bot VPN  

Защита и свобода в сети —  
невидимость, доступ, скорость.  
Выбери функцию:
""",
                    reply_markup=main_menu_keyboard()
                )
    except Exception as e:
        logger.error(f"Ошибка в start_handler: {e}")
        
@bot.callback_query_handler(func=lambda call: call.data == "check_sub")
def check_subscription_callback(call):
    """Обработчик кнопки проверки подписки"""
    try:
        user_id = call.from_user.id
        
        if check_subscription(user_id):
            bot.delete_message(call.message.chat.id, call.message.message_id)
            
            # После проверки показываем стандартное приветствие
            if is_admin(user_id):
                bot.send_message(call.message.chat.id, "👋 Добро пожаловать в админ-панель!", reply_markup=admin_menu_keyboard())
            else:
                try:
                    with open('welcome.png', 'rb') as photo:
                        bot.send_photo(
                            call.message.chat.id, 
                            photo,
                            caption="""💙 Добро пожаловать в you name bot VPN  

Защита и свобода в сети —  
невидимость, доступ, скорость.  
Выбери функцию:
""",
                            reply_markup=main_menu_keyboard()
                        )
                except FileNotFoundError:
                    bot.send_message(
                        call.message.chat.id,
                        """💙 Добро пожаловать в you name bot VPN  

Защита и свобода в сети —  
невидимость, доступ, скорость.  
Выбери функцию:
""",
                        reply_markup=main_menu_keyboard()
                    )
        else:
            bot.answer_callback_query(
                callback_query_id=call.id,
                text="❌ Вы подписались не на все каналы!",
                show_alert=True
            )
    except Exception as e:
        logger.error(f"Ошибка в check_subscription_callback: {e}")
        
# Декоратор для проверки подписки перед любыми действиями (кроме админов и start)
@bot.message_handler(func=lambda message: 
                     not is_admin(message.from_user.id) and 
                     not check_subscription(message.from_user.id) and
                     message.text not in ['/start', '/help'])
def require_subscription_wrapper(message):
    """Требует подписку на каналы для доступа к функциям бота"""
    try:
        user_id = message.from_user.id
        
        bot.send_message(
            user_id,
            "👋 Для использования бота необходимо подписаться на наши каналы:",
            reply_markup=subscription_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка в require_subscription_wrapper: {e}")                      

@bot.message_handler(func=lambda message: message.text == '🔙 В главное меню' and is_admin(message.from_user.id))
def back_to_main_admin(message):
    start_handler(message)

@bot.message_handler(func=lambda message: message.text == '🔙 Назад')
def back_handler(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        if is_admin(user_id):
            bot.send_message(user_id, "Возвращаемся в админ-меню", reply_markup=admin_menu_keyboard())
        else:
            bot.send_message(user_id, "Возвращаемся в главное меню", reply_markup=main_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в back_handler: {e}")


# Обработчик выбора сервера Нидерланды
@bot.message_handler(func=lambda message: message.text == '🇳🇱 Netherlands')
def select_netherlands_server(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        server_name = '🇳🇱 Netherlands'
        
        user_data = get_user_data(user_id)
        user_data['selected_server'] = server_name
        save_user_data(user_id, user_data)
        
        bot.send_message(user_id, f"Выбрали сервер: {server_name}\nВыберите срок:", reply_markup=duration_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в select_netherlands_server: {e}")
        
@bot.message_handler(func=lambda message: message.text == '🛒 Купить VPN | Продлить')
def buy_or_extend_vpn(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        # Проверяем есть ли доступные конфиги на серверах
        available_configs = False
        for server in servers_db.values():
            if server['available_configs']:
                available_configs = True
                break
        
        if not available_configs:
            bot.send_message(
                user_id,
                "❌ <b>В настоящее время нет доступных конфигурационных файлов.</b>\n\n"
                "Администратор уже уведомлен и скоро добавит новые конфиги.\n"
                "Пожалуйста, попробуйте позже.",
                parse_mode='HTML',
                reply_markup=main_menu_keyboard()
            )
            
            # Уведомляем администратора
            bot.send_message(
                ADMIN_ID,
                "⚠️ <b>ВНИМАНИЕ: ЗАКОНЧИЛИСЬ КОНФИГИ!</b>\n\n"
                "Пользователь пытался купить VPN, но конфиги закончились.\n"
                "Срочно добавьте новые конфигурационные файлы!",
                parse_mode='HTML'
            )
            return
        
        # Проверяем есть ли у пользователя активные подписки
        user_data = get_user_data(user_id)
        active_subs = [sub for sub in user_data.get('subscriptions', []) 
                      if datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S") > datetime.now()]
        
        if active_subs:
            # Создаем клавиатуру с кнопками продления и покупки нового
            keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            for sub in active_subs:
                keyboard.add(types.KeyboardButton(f'🔄 Продлить {sub["server"]} ({sub["config_file"]})'))
            keyboard.add(types.KeyboardButton('➕ Купить новый ключ'))
            keyboard.add(types.KeyboardButton('🔙 Назад'))
            
            bot.send_message(
                user_id,
                "У вас есть активные подписки. Вы можете продлить существующие или купить новый ключ:",
                reply_markup=keyboard
            )
        else:
            # Предлагаем купить
            bot.send_message(user_id, "Выберите сервер:", reply_markup=servers_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в buy_or_extend_vpn: {e}")

# Добавляем обработчик для покупки нового ключа
@bot.message_handler(func=lambda message: message.text == '➕ Купить новый ключ')
def buy_new_key(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        bot.send_message(user_id, "Выберите сервер для нового ключа:", reply_markup=servers_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в buy_new_key: {e}")

@bot.message_handler(func=lambda message: message.text.startswith('🔄 Продлить '))
def handle_extend_subscription(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        # Извлекаем данные из кнопки
        button_text = message.text
        server_name = button_text.replace('🔄 Продлить ', '').split(' (')[0]
        config_file = button_text.split('(')[1].replace(')', '')
        
        # Проверяем доступность конфигов на сервере
        server_key = next((k for k, v in servers_db.items() if v['name'] == server_name), None)
        if not server_key or not servers_db[server_key]['available_configs']:
            bot.send_message(
                user_id,
                f"❌ <b>На сервере {server_name} временно нет доступных конфигов.</b>\n\n"
                "Администратор уже уведомлен и скоро добавит новые.\n"
                "Попробуйте позже или выберите другой сервер.",
                parse_mode='HTML',
                reply_markup=main_menu_keyboard()
            )
            
            # Уведомляем администратора
            bot.send_message(
                ADMIN_ID,
                f"⚠️ <b>ВНИМАНИЕ: ЗАКОНЧИЛИСЬ КОНФИГИ НА СЕРВЕРЕ {server_name}!</b>\n\n"
                f"Пользователь @{message.from_user.username} пытался продлить подписку.\n"
                "Срочно добавьте новые конфигурационные файлы!",
                parse_mode='HTML'
            )
            return
        
        # Сохраняем выбранный сервер и конфиг для продления
        user_data = get_user_data(user_id)
        user_data['selected_server'] = server_name
        user_data['selected_config'] = config_file
        save_user_data(user_id, user_data)
        
        # Предлагаем выбрать срок продления
        bot.send_message(
            user_id,
            f"Вы выбрали продление подписки для сервера {server_name}.\nВыберите срок продления:",
            reply_markup=duration_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка в handle_extend_subscription: {e}")
        bot.send_message(user_id, "Произошла ошибка при обработке запроса.")
        
@bot.message_handler(func=lambda message: message.text == '📤 Добавить конфиги массово' and is_admin(message.from_user.id))
def bulk_upload_configs(message):
    """Массовая загрузка конфигурационных файлов"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Создаем клавиатуру с выбором сервера
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        for server in servers_db.values():
            keyboard.add(types.KeyboardButton(f"📁 Массовая загрузка {server['name']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "🖥 <b>Массовая загрузка конфигов</b>\n\n"
            "Выберите сервер для загрузки конфигурационных файлов.\n\n"
            "💡 <b>Как использовать:</b>\n"
            "1. Выберите сервер\n"  
            "2. Отправьте один или несколько .conf файлов\n"
            "3. Файлы автоматически добавятся на выбранный сервер",
            parse_mode='HTML',
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_bulk_server_selection)
    except Exception as e:
        logger.error(f"Ошибка в bulk_upload_configs: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при запуске массовой загрузки")

def process_bulk_server_selection(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        # Извлекаем название сервера из кнопки
        server_name = message.text.replace('📁 Массовая загрузка ', '')
        
        # Проверяем существование сервера
        server_key = next((k for k, v in servers_db.items() if v['name'] == server_name), None)
        if not server_key:
            bot.send_message(message.chat.id, "❌ Сервер не найден", reply_markup=admin_menu_keyboard())
            return
        
        # Сохраняем выбранный сервер во временных данных
        user_data = get_user_data(message.from_user.id)
        user_data['bulk_upload_server'] = server_name
        user_data['bulk_upload_files'] = []  # Список для отслеживания загруженных файлов
        save_user_data(message.from_user.id, user_data)
        
        # Показываем текущую статистику сервера
        server_data = servers_db[server_key]
        stats_text = (
            f"📤 <b>Массовая загрузка для {server_name}</b>\n\n"
            f"📊 <b>Текущая статистика сервера:</b>\n"
            f"🆓 Доступно конфигов: {len(server_data['available_configs'])}\n"
            f"👥 Используется: {len(server_data['used_configs'])}\n\n"
            f"📨 <b>Теперь отправьте .conf файлы</b>\n\n"
            f"💡 <b>Совет:</b> Можно отправить несколько файлов за раз\n"
            f"📝 После загрузки файлов нажмите /done для завершения"
        )
        
        msg = bot.send_message(
            message.chat.id,
            stats_text,
            parse_mode='HTML',
            reply_markup=types.ReplyKeyboardRemove()
        )
        
        # Регистрируем обработчик для документов
        bot.register_next_step_handler(msg, process_bulk_configs_upload)
        
    except Exception as e:
        logger.error(f"Ошибка в process_bulk_server_selection: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка выбора сервера", reply_markup=admin_menu_keyboard())

def process_bulk_configs_upload(message):
    try:
        user_id = message.from_user.id
        user_data = get_user_data(user_id)
        server_name = user_data.get('bulk_upload_server')
        
        if not server_name:
            bot.send_message(message.chat.id, "❌ Ошибка: сервер не выбран", reply_markup=admin_menu_keyboard())
            return
        
        # Проверяем команду завершения
        if message.text and message.text == '/done':
            uploaded_files = user_data.get('bulk_upload_files', [])
            send_bulk_upload_summary(message.chat.id, server_name, uploaded_files)
            # Очищаем временные данные
            if 'bulk_upload_server' in user_data:
                del user_data['bulk_upload_server']
            if 'bulk_upload_files' in user_data:
                del user_data['bulk_upload_files']
            save_user_data(user_id, user_data)
            return
        
        if message.content_type != 'document':
            bot.send_message(
                message.chat.id, 
                "❌ Пожалуйста, отправьте .conf файлы или введите /done для завершения",
                reply_markup=admin_menu_keyboard()
            )
            return
        
        # Обрабатываем файл
        success = process_single_config_file(message.document, server_name, user_id)
        
        if success:
            # Добавляем файл в список загруженных
            if 'bulk_upload_files' not in user_data:
                user_data['bulk_upload_files'] = []
            user_data['bulk_upload_files'].append(message.document.file_name)
            save_user_data(user_id, user_data)
            
            # Отправляем подтверждение
            bot.send_message(
                message.chat.id,
                f"✅ Файл '{message.document.file_name}' успешно добавлен!\n"
                f"Отправьте следующий файл или введите /done для завершения",
                reply_markup=types.ReplyKeyboardRemove()
            )
            
            # Регистрируем следующий шаг
            bot.register_next_step_handler(message, process_bulk_configs_upload)
        else:
            bot.send_message(
                message.chat.id,
                f"❌ Ошибка при добавлении файла '{message.document.file_name}'\n"
                f"Убедитесь что это .conf файл и попробуйте снова",
                reply_markup=types.ReplyKeyboardRemove()
            )
            bot.register_next_step_handler(message, process_bulk_configs_upload)
        
    except Exception as e:
        logger.error(f"Ошибка в process_bulk_configs_upload: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при загрузке файлов", reply_markup=admin_menu_keyboard())
        
def send_bulk_upload_summary(chat_id, server_name, uploaded_files):
    """Отправляет сводку о массовой загрузке"""
    try:
        text = f"📊 <b>Сводка массовой загрузки для {server_name}</b>\n\n"
        text += f"✅ Успешно загружено: {len(uploaded_files)} файлов\n"
        text += f"📁 Всего обработано: {len(uploaded_files)} файлов\n\n"
        
        if uploaded_files:
            text += "<b>Загруженные файлы:</b>\n"
            for file in uploaded_files[:10]:  # Показываем первые 10 файлов
                text += f"• {file}\n"
            
            if len(uploaded_files) > 10:
                text += f"... и еще {len(uploaded_files) - 10} файлов\n"
        
        # Обновляем статистику сервера
        server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
        available_count = len(servers_db[server_key]['available_configs'])
        
        text += f"\n📈 <b>Текущая статистика сервера:</b>\n"
        text += f"🆓 Доступно конфигов: {available_count}\n"
        text += f"👥 Используется: {len(servers_db[server_key]['used_configs'])}\n"
        
        bot.send_message(chat_id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка отправки сводки: {e}")
        
@bot.message_handler(commands=['done'])
def handle_done_command(message):
    """Обрабатывает команду завершения массовой загрузки"""
    if not is_admin(message.from_user.id):
        return
    
    user_data = get_user_data(message.from_user.id)
    server_name = user_data.get('bulk_upload_server')
    
    if server_name:
        uploaded_files = user_data.get('bulk_upload_files', [])
        send_bulk_upload_summary(message.chat.id, server_name, uploaded_files)
        
        # Очищаем временные данные
        if 'bulk_upload_server' in user_data:
            del user_data['bulk_upload_server']
        if 'bulk_upload_files' in user_data:
            del user_data['bulk_upload_files']
        save_user_data(message.from_user.id, user_data)
    else:
        bot.send_message(message.chat.id, "❌ Нет активной массовой загрузки", reply_markup=admin_menu_keyboard())

def process_single_config_file(document, server_name, user_id):
    """Обрабатывает один конфигурационный файл"""
    try:
        # Проверяем что файл имеет правильное расширение
        if not document.file_name.endswith('.conf'):
            return False
        
        file_info = bot.get_file(document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        # Сохраняем файл с оригинальным именем
        config_filename = document.file_name
        
        with open(config_filename, 'wb') as new_file:
            new_file.write(downloaded_file)
        
        # Добавляем конфиг в доступные для сервера
        server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
        if config_filename not in servers_db[server_key]['available_configs']:
            servers_db[server_key]['available_configs'].append(config_filename)
        
        save_data_to_file()
        
        logger.info(f"Админ {user_id} добавил конфиг {config_filename} для сервера {server_name}")
        return True
        
    except Exception as e:
        logger.error(f"Ошибка обработки файла {document.file_name}: {e}")
        return False

# Добавляем обработчик для группы медиа (несколько файлов)
@bot.message_handler(content_types=['document'], func=lambda message: hasattr(message, 'media_group_id'))
def handle_media_group(message):
    """Обрабатывает группу файлов (несколько файлов отправленных вместе)"""
    try:
        # Эта функция будет вызываться для каждого файла в группе
        user_id = message.from_user.id
        if not is_admin(user_id):
            return
        
        user_data = get_user_data(user_id)
        server_name = user_data.get('bulk_upload_server')
        
        if server_name and message.document and message.document.file_name.endswith('.conf'):
            success = process_single_config_file(message.document, server_name, user_id)
            if success:
                logger.info(f"Успешно добавлен конфиг из группы: {message.document.file_name}")
    
    except Exception as e:
        logger.error(f"Ошибка в handle_media_group: {e}")
        
def send_bulk_upload_summary(chat_id, server_name, uploaded_files, total_files):
    """Отправляет сводку о массовой загрузке"""
    try:
        text = f"📊 <b>Сводка массовой загрузки для {server_name}</b>\n\n"
        text += f"✅ Успешно загружено: {len(uploaded_files)} файлов\n"
        text += f"📁 Всего обработано: {total_files} файлов\n\n"
        
        if uploaded_files:
            text += "<b>Загруженные файлы:</b>\n"
            for file in uploaded_files[:10]:  # Показываем первые 10 файлов
                text += f"• {file}\n"
            
            if len(uploaded_files) > 10:
                text += f"... и еще {len(uploaded_files) - 10} файлов\n"
        
        # Обновляем статистику сервера
        server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
        available_count = len(servers_db[server_key]['available_configs'])
        
        text += f"\n📈 <b>Текущая статистика сервера:</b>\n"
        text += f"🆓 Доступно конфигов: {available_count}\n"
        text += f"👥 Используется: {len(servers_db[server_key]['used_configs'])}\n"
        
        bot.send_message(chat_id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка отправки сводки: {e}")

@bot.message_handler(func=lambda message: message.text == '🇩🇪 Germany')
def select_server(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        server_name = '🇩🇪 Germany'
        
        user_data = get_user_data(user_id)
        user_data['selected_server'] = server_name
        save_user_data(user_id, user_data)
        
        bot.send_message(user_id, f"Выбрали сервер: {server_name}\nВыберите срок:", reply_markup=duration_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в select_server: {e}")

@bot.message_handler(func=lambda message: any(duration in message.text for duration in ['1 месяц', '3 месяца', '6 месяцев']))
def select_duration(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        duration = message.text.split(' - ')[0]
        price = message.text.split(' - ')[1]
        
        user_data = get_user_data(user_id)
        if 'selected_server' not in user_data:
            bot.send_message(user_id, "Пожалуйста, сначала выберите сервер.")
            return
            
        user_data['selected_duration'] = duration
        user_data['selected_price'] = price
        save_user_data(user_id, user_data)
        
        # Отправляем изображение selectbank.png перед выбором банка
        try:
            with open('selectbank.png', 'rb') as photo:
                bot.send_photo(
                    user_id,
                    photo,
                    caption=f"""
Сервер: {user_data['selected_server']}
Срок: {duration}
Сумма: {price}

Выберите банк:
""",
                    reply_markup=payment_methods_keyboard()
                )
        except FileNotFoundError:
            bot.send_message(
                user_id, 
                f"""
Сервер: {user_data['selected_server']}
Срок: {duration}
Сумма: {price}

Выберите банк:
""",
                reply_markup=payment_methods_keyboard()
            )
    except Exception as e:
        logger.error(f"Ошибка в select_duration: {e}")

@bot.message_handler(func=lambda message: any(method['bank'] in message.text for method in payment_methods.values()))
def select_payment_method(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        bank_name = message.text
        user_data = get_user_data(user_id)
        server_name = user_data['selected_server']
        
        method_key = next(key for key, val in payment_methods.items() if val['bank'] == bank_name)
        method = payment_methods[method_key]
        
        payment_id = generate_payment_id()
        
        payment_data = {
            'user_id': user_id,
            'username': message.from_user.username,
            'server': server_name,
            'duration': user_data['selected_duration'],
            'amount': user_data['selected_price'],
            'bank': method['bank'],
            'status': 'pending',
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        save_payment(payment_id, payment_data)
        
        # Клавиатура с подтверждением оплаты
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('✅ Я оплатил'))
        markup.add(types.KeyboardButton('🔙 Назад'))
        
        # Определяем какое фото показывать на основе названия банка (независимо от регистра)
        bank_name_lower = bank_name.lower()
        
        # Определяем имя файла изображения
        photo_file = None
        
        # Проверяем различные варианты написания банков
        if 'сбер' in bank_name_lower or 'sber' in bank_name_lower:
            photo_file = 'sber.png'
        elif 'альфа' in bank_name_lower or 'alfa' in bank_name_lower:
            photo_file = 'alfa.png'
        elif 'тинькофф' in bank_name_lower or 'тиньков' in bank_name_lower or 'tinkoff' in bank_name_lower:
            photo_file = 'tbank.PNG'
        elif 'тбанк' in bank_name_lower or 'tbank' in bank_name_lower:
            photo_file = 'tbank.PNG'
        else:
            # Если нет специального изображения, используем стандартное
            photo_file = 'selectbank.png'
        
        # Формируем текст сообщения в зависимости от банка
        if 'сбер' in bank_name_lower or 'sber' in bank_name_lower:
            message_text = f"""
💳 СБЕР ОПЛАТА
Номер: <code>{method['card_number']}</code>
Сумма: {user_data['selected_price']}

⚠️ В комментарии укажите код: <code>{payment_id}</code>

После перевода нажмите ✅ Я оплатил
"""
        else:
            message_text = f"""
💳 Реквизиты для оплаты ({method['bank']}):

Номер для перевода: <code>{method['card_number']}</code>
Сумма: {user_data['selected_price']}
⚠️ В комментарии перевода обязательно укажите код: <code>{payment_id}</code>

После перевода нажмите ✅ Я оплатил
"""
        
        # Пытаемся отправить фото, если файл существует
        try:
            if os.path.exists(photo_file):
                with open(photo_file, 'rb') as photo:
                    bot.send_photo(
                        user_id,
                        photo,
                        caption=message_text,
                        parse_mode='HTML',
                        reply_markup=markup
                    )
            else:
                # Если файл не найден, используем стандартное изображение или текст
                if photo_file != 'selectbank.png' and os.path.exists('selectbank.png'):
                    with open('selectbank.png', 'rb') as photo:
                        bot.send_photo(
                            user_id,
                            photo,
                            caption=message_text,
                            parse_mode='HTML',
                            reply_markup=markup
                        )
                else:
                    bot.send_message(
                        user_id,
                        message_text,
                        parse_mode='HTML',
                        reply_markup=markup
                    )
        except Exception as photo_error:
            logger.error(f"Ошибка при отправке фото: {photo_error}")
            # Если ошибка с фото, отправляем просто текст
            bot.send_message(
                user_id,
                message_text,
                parse_mode='HTML',
                reply_markup=markup
            )
            
    except Exception as e:
        logger.error(f"Ошибка в select_payment_method: {e}")
        
@bot.message_handler(func=lambda message: message.text == '🎁 Управление промокодами' and is_admin(message.from_user.id))
def manage_promo_codes(message):
    try:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('➕ Добавить промокод'))
        markup.add(types.KeyboardButton('➖ Удалить промокод'))
        markup.add(types.KeyboardButton('📋 Список промокодов'))
        markup.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(
            message.chat.id,
            "Управление промокодами:",
            reply_markup=markup
        )
    except Exception as e:
        logger.error(f"Ошибка в manage_promo_codes: {e}")

@bot.message_handler(func=lambda message: message.text == '➕ Добавить промокод' and is_admin(message.from_user.id))
def add_promo_code(message):
    try:
        # Создаем клавиатуру с доступными серверами
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        for server in servers_db.values():
            keyboard.add(types.KeyboardButton(f"🎁 {server['name']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "Выберите сервер для промокода:",
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_promo_server_selection)
    except Exception as e:
        logger.error(f"Ошибка в add_promo_code: {e}")
        
def process_promo_server_selection(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        server_name = message.text.replace('🎁 ', '')
        
        msg = bot.send_message(
            message.chat.id,
            f"Введите данные промокода для {server_name} в формате:\nКод: GERMANY21\nДней: 21",
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, lambda m: process_add_promo(m, server_name))
    except Exception as e:
        logger.error(f"Ошибка в process_promo_server_selection: {e}")

def process_add_promo(message, server_name):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        lines = [line.strip() for line in message.text.split('\n') if line.strip()]
        promo_data = {}
        
        for line in lines:
            if line.startswith('Код:'):
                promo_data['code'] = line.split('Код:')[1].strip()
            elif line.startswith('Дней:'):
                promo_data['days'] = int(line.split('Дней:')[1].strip())
        
        # Устанавливаем выбранный сервер
        promo_data['server'] = server_name
        
        if not all(key in promo_data for key in ['code', 'days']):
            raise ValueError("Неполные данные")
        
        PROMO_CODES[promo_data['code']] = {
            'server': promo_data['server'],
            'days': promo_data['days'],
            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'created_by': message.from_user.id
        }
        
        with open('promo_codes.json', 'w') as f:
            json.dump(PROMO_CODES, f)
            
        bot.send_message(
            message.chat.id,
            f"✅ Промокод {promo_data['code']} добавлен!\n"
            f"Сервер: {promo_data['server']}\n"
            f"Дней: {promo_data['days']}",
            reply_markup=admin_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка добавления промокода: {e}")
        bot.send_message(message.chat.id, "Ошибка формата данных. Используйте формат:\nКод: SERVER21\nДней: 21")

@bot.message_handler(func=lambda message: message.text == '➖ Удалить промокод' and is_admin(message.from_user.id))
def delete_promo_code(message):
    try:
        if not PROMO_CODES:
            bot.send_message(message.chat.id, "Нет промокодов для удаления.")
            return
            
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        for code in PROMO_CODES.keys():
            markup.add(types.KeyboardButton(code))
        markup.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "Выберите промокод для удаления:",
            reply_markup=markup
        )
        bot.register_next_step_handler(msg, process_delete_promo)
    except Exception as e:
        logger.error(f"Ошибка в delete_promo_code: {e}")

def process_delete_promo(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        code = message.text.strip()
        if code in PROMO_CODES:
            del PROMO_CODES[code]
            with open('promo_codes.json', 'w') as f:
                json.dump(PROMO_CODES, f)
            bot.send_message(
                message.chat.id,
                f"Промокод {code} успешно удален!",
                reply_markup=admin_menu_keyboard()
            )
        else:
            bot.send_message(message.chat.id, "Промокод не найден.")
    except Exception as e:
        logger.error(f"Ошибка в process_delete_promo: {e}")

# Добавляем обработчики для новых кнопок
@bot.message_handler(func=lambda message: message.text == '💾 Создать резервную копию' and is_admin(message.from_user.id))
def backup_data(message):
    try:
        backup_file = create_backup()
        if backup_file:
            with open(backup_file, 'rb') as f:
                bot.send_document(
                    message.chat.id,
                    f,
                    caption="✅ Резервная копия создана успешно!",
                    visible_file_name=os.path.basename(backup_file)
                )
        else:
            bot.send_message(message.chat.id, "❌ Ошибка создания резервной копии")
    except Exception as e:
        logger.error(f"Ошибка в backup_data: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка создания резервной копии")

@bot.message_handler(func=lambda message: message.text == '📥 Восстановить из копии' and is_admin(message.from_user.id))
def restore_data(message):
    try:
        msg = bot.send_message(
            message.chat.id,
            "📤 Отправьте файл резервной копии (ZIP архив):",
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, process_restore_file)
    except Exception as e:
        logger.error(f"Ошибка в restore_data: {e}")
        
def process_restore_file(message):
    try:
        if message.content_type != 'document':
            bot.send_message(message.chat.id, "❌ Пожалуйста, отправьте ZIP файл")
            return
        
        if not message.document.file_name.endswith('.zip'):
            bot.send_message(message.chat.id, "❌ Файл должен быть в формате ZIP")
            return
        
        # Скачиваем файл
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        # Сохраняем временный файл
        temp_file = 'temp_backup.zip'
        with open(temp_file, 'wb') as f:
            f.write(downloaded_file)
        
        # Восстанавливаем из резервной копии
        success = restore_from_backup(temp_file)
        
        # Удаляем временный файл
        os.remove(temp_file)
        
        if success:
            bot.send_message(
                message.chat.id,
                "✅ Данные успешно восстановлены из резервной копии!\n"
                "Все существующие данные были объединены с резервной копией.",
                reply_markup=admin_menu_keyboard()
            )
        else:
            bot.send_message(
                message.chat.id,
                "❌ Ошибка восстановления данных",
                reply_markup=admin_menu_keyboard()
            )
            
    except Exception as e:
        logger.error(f"Ошибка в process_restore_file: {e}")
        bot.send_message(
            message.chat.id,
            "❌ Ошибка обработки файла",
            reply_markup=admin_menu_keyboard()
        )

@bot.message_handler(func=lambda message: message.text == '📋 Список промокодов' and is_admin(message.from_user.id))
def list_promo_codes(message):
    try:
        if not PROMO_CODES:
            bot.send_message(message.chat.id, "Нет активных промокодов.")
            return
            
        text = "📋 Список промокодов:\n\n"
        for code, data in PROMO_CODES.items():
            text += f"🔹 Код: <code>{code}</code>\n"
            text += f"🌍 Сервер: {data['server']}\n"
            text += f"⏳ Дней: {data['days']}\n"
            text += f"📅 Создан: {data['created_at']}\n\n"
        
        bot.send_message(
            message.chat.id,
            text,
            parse_mode='HTML',
            reply_markup=admin_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка в list_promo_codes: {e}")
        
@bot.message_handler(func=lambda message: message.text == '📲 Установить приложение')
def install_app(message):
    markup = types.InlineKeyboardMarkup()
    markup.row(
        types.InlineKeyboardButton("Android", url="https://play.google.com/store/apps/details?id=org.amnezia.awg"),
        types.InlineKeyboardButton("iOS", url="https://apps.apple.com/ru/app/amneziawg/id6478942365")
    )
    markup.row(
        types.InlineKeyboardButton("Windows", url="https://github.com/amnezia-vpn/amneziawg-windows-client/releases/tag/1.0.2"),
        types.InlineKeyboardButton("macOS", url="https://apps.apple.com/us/app/amneziawg/id6478942365")
    )
    
    bot.send_message(
        message.chat.id,
        "📲 Скачайте AmneziaWG для вашей платформы:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: message.text == 'ℹ️ История покупок')
def purchase_history(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        user_data = get_user_data(user_id)
        if 'subscriptions' not in user_data or not user_data['subscriptions']:
            bot.send_message(user_id, "У вас пока нет завершенных покупок.")
            return
        
        text = "📋 История ваших покупок:\n\n"
        for idx, sub in enumerate(reversed(user_data['subscriptions']), 1):
            expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
            is_expired = expiry_date < datetime.now()
            
            text += f"{idx}. Сервер: {sub['server']}\n"
            text += f"Файл: {sub['config_file']}\n"
            text += f"Срок: {sub.get('duration', 'N/A')}\n"
            text += f"Дата покупки: {sub['purchase_date']}\n"
            text += f"Действует до: {sub['expiry_date']}"
            
            if is_expired:
                text += " (⚠️ Истек)\n\n"
            else:
                text += " (✅ Активен)\n\n"
            
            # Отправляем файл конфига
            if os.path.exists(sub['config_file']):
                with open(sub['config_file'], 'rb') as f:
                    bot.send_document(
                        user_id,
                        f,
                        caption=f"Конфигурация для покупки #{idx}",
                        visible_file_name=os.path.basename(sub['config_file'])
                    )
            else:
                bot.send_message(user_id, f"Файл конфигурации {sub['config_file']} не найден.")
        
        bot.send_message(user_id, text)
    except Exception as e:
        logger.error(f"Ошибка в purchase_history: {e}")

@bot.message_handler(func=lambda message: message.text == '🛟 Поддержка')
def support(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        bot.send_message(message.chat.id, "По всем вопросам обращайтесь к @you username")
    except Exception as e:
        logger.error(f"Ошибка в support: {e}")

@bot.message_handler(func=lambda message: message.text == '⭐⭐⭐ Отзывы')
def show_reviews(message):
    try:
        markup = types.InlineKeyboardMarkup()
        btn_reviews = types.InlineKeyboardButton("📢 Посмотреть отзывы", url="https://t.me/you name bot_09/28")
        markup.add(btn_reviews)
        bot.send_message(
            message.chat.id,
            "Отзывы реальных людей у нас на канале:",
            reply_markup=markup
        )
    except Exception as e:
        logger.error(f"Ошибка в show_reviews: {e}")
        
# Админ-обработчики
@bot.message_handler(func=lambda message: message.text == '📊 Статистика' and is_admin(message.from_user.id))
def stats(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        total_users = len(users_db)
        total_payments = len(payments_db)
        approved_payments = len([p for p in payments_db.values() if p.get('status') == 'approved'])
        pending_payments = len([p for p in payments_db.values() if p.get('status') == 'pending'])
        
        # Правильный расчет дохода
        revenue = 0
        for p in payments_db.values():
            if p.get('status') == 'approved':
                amount_str = p.get('amount', '0₽')
                # Убираем символ валюты и преобразуем в число
                amount_num = int(''.join(filter(str.isdigit, amount_str)))
                revenue += amount_num
        
        # Подсчет пользователей с бесплатной подпиской, которые потом купили
        converted_users = 0
        
        # Подсчет продлений - ИСПРАВЛЕННАЯ ЛОГИКА
        extended_count = 0
        users_with_extensions = set()
        config_extensions = {}
        
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                # Считаем количество платежей для каждого конфига
                config_payments = {}
                for sub in user_data['subscriptions']:
                    if 'payment_id' in sub:  # Только платные подписки
                        config_file = sub['config_file']
                        if config_file not in config_payments:
                            config_payments[config_file] = 0
                        config_payments[config_file] += 1
                
                # Если у конфига больше 1 платежа - это продление
                for config_file, payment_count in config_payments.items():
                    if payment_count > 1:
                        extensions = payment_count - 1
                        extended_count += extensions
                        users_with_extensions.add(user_id)
                        
                        if config_file not in config_extensions:
                            config_extensions[config_file] = 0
                        config_extensions[config_file] += extensions
        
        # Формируем детальную информацию о продлениях
        extensions_text = ""
        if extended_count > 0:
            extensions_text = f"\n🔁 <b>Продления:</b> {extended_count}\n"
            extensions_text += f"👥 <b>Пользователей с продлениями:</b> {len(users_with_extensions)}\n\n"
            
            # Топ продлеваемых конфигов
            sorted_configs = sorted(config_extensions.items(), key=lambda x: x[1], reverse=True)[:5]
            extensions_text += "<b>Топ продлеваемых конфигов:</b>\n"
            
            for config_file, extensions_count in sorted_configs:
                short_name = os.path.basename(config_file)
                if len(short_name) > 20:
                    short_name = short_name[:17] + "..."
                extensions_text += f"• {short_name}: {extensions_count} продл.\n"
        else:
            extensions_text = f"\n🔁 <b>Продления:</b> 0\n👥 <b>Пользователей с продлениями:</b> 0\n"
        
        bot.send_message(message.chat.id, f"""
📊 <b>Статистика:</b>
👥 <b>Пользователей:</b> {total_users}
💳 <b>Платежей:</b> {total_payments}
✅ <b>Подтверждено:</b> {approved_payments}
⏳ <b>Ожидает:</b> {pending_payments}
💰 <b>Доход:</b> {revenue}₽
🔄 <b>Конверсия (промо → покупка):</b> {converted_users} пользователей
{extensions_text}
""", parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка в stats: {e}")
        
@bot.message_handler(func=lambda message: message.text == '🔁 Статистика продлений' and is_admin(message.from_user.id))
def extensions_stats(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        extended_configs = {}
        user_extensions = {}
        
        # Собираем статистику по продлениям
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                config_payments = {}
                
                for sub in user_data['subscriptions']:
                    if 'payment_id' in sub:  # Только платные подписки
                        config_file = sub['config_file']
                        if config_file not in config_payments:
                            config_payments[config_file] = 0
                        config_payments[config_file] += 1
                
                # Считаем продления для каждого конфига
                for config_file, payment_count in config_payments.items():
                    if payment_count > 1:  # Если больше 1 платежа - есть продления
                        extensions_count = payment_count - 1
                        
                        if config_file not in extended_configs:
                            extended_configs[config_file] = 0
                        extended_configs[config_file] += extensions_count
                        
                        if user_id not in user_extensions:
                            user_extensions[user_id] = {
                                'username': user_data.get('username', 'N/A'),
                                'total_extensions': 0,
                                'configs': {}
                            }
                        user_extensions[user_id]['total_extensions'] += extensions_count
                        user_extensions[user_id]['configs'][config_file] = extensions_count
        
        if not extended_configs:
            bot.send_message(message.chat.id, "❌ Нет данных о продлениях.", reply_markup=admin_menu_keyboard())
            return
        
        # Формируем подробный отчет
        text = "📈 <b>Детальная статистика продлений</b>\n\n"
        
        # Общая информация
        total_extensions = sum(extended_configs.values())
        total_users_with_extensions = len(user_extensions)
        
        text += f"🔁 <b>Всего продлений:</b> {total_extensions}\n"
        text += f"👥 <b>Пользователей с продлениями:</b> {total_users_with_extensions}\n"
        text += f"🖥 <b>Продлеваемых конфигов:</b> {len(extended_configs)}\n\n"
        
        # Топ конфигов по продлениям
        text += "🏆 <b>Топ конфигов по продлениям:</b>\n"
        sorted_configs = sorted(extended_configs.items(), key=lambda x: x[1], reverse=True)[:10]
        
        for i, (config_file, extensions_count) in enumerate(sorted_configs, 1):
            short_name = os.path.basename(config_file)
            if len(short_name) > 25:
                short_name = short_name[:22] + "..."
            text += f"{i}. {short_name}: {extensions_count} продл.\n"
        
        text += "\n🏆 <b>Топ пользователей по продлениям:</b>\n"
        sorted_users = sorted(user_extensions.items(), key=lambda x: x[1]['total_extensions'], reverse=True)[:5]
        
        for i, (user_id, data) in enumerate(sorted_users, 1):
            text += f"{i}. @{data['username']}: {data['total_extensions']} продл.\n"
            for config_file, ext_count in data['configs'].items():
                short_name = os.path.basename(config_file)
                if len(short_name) > 15:
                    short_name = short_name[:12] + "..."
                text += f"   • {short_name}: {ext_count}\n"
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка в extensions_stats: {e}")
        

@bot.message_handler(func=lambda message: message.text == '📝 Список серверов' and is_admin(message.from_user.id))
def manage_servers(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Получаем все конфиги из файловой системы
        all_configs = [f for f in os.listdir() if f.endswith('.conf')]
        
        text = "🖥 Реальные серверы:\n\n"
        
        # Пересчитываем статистику для каждого сервера
        for server_key, server_data in servers_db.items():
            server_name = server_data['name']
            
            # Определяем паттерн для поиска конфигов этого сервера
            if 'Germany' in server_name or 'DE' in server_name or '🇩🇪' in server_name:
                patterns = ['Grm', 'germany', 'Germany', 'DE']
            elif 'Netherlands' in server_name or 'NL' in server_name or '🇳🇱' in server_name:
                patterns = ['Ndr', 'netherlands', 'Netherlands', 'NL']
            else:
                patterns = [server_name.replace(' ', '_')]
            
            # Ищем конфиги по паттернам
            server_configs = []
            for config in all_configs:
                for pattern in patterns:
                    if pattern in config:
                        server_configs.append(config)
                        break
            
            # Получаем актуальные данные из базы
            used_configs_count = len(server_data['used_configs'])
            available_configs_count = len(server_data['available_configs'])
            total_configs_count = used_configs_count + available_configs_count
            
            text += f"<b>{server_name}</b>\n"
            text += f"📍 Локация: {server_data['location']}\n"
            text += f"🖥 IP: {server_data['ip']}\n"
            text += f"⚡ Нагрузка: {server_data['load']}\n"
            text += f"📊 Всего конфигов: {total_configs_count}\n"
            text += f"📈 Доступно конфигов: {available_configs_count}\n"
            text += f"👥 Используется: {used_configs_count}\n"
            
            # Показываем несколько примеров конфигов
            if server_data['available_configs']:
                sample_configs = server_data['available_configs'][:3]
                text += f"📁 Примеры доступных: {', '.join(sample_configs)}"
                if len(server_data['available_configs']) > 3:
                    text += f" ... и еще {len(server_data['available_configs']) - 3}\n"
                else:
                    text += "\n"
            
            if server_data['used_configs']:
                sample_used = list(server_data['used_configs'].values())[:2]
                text += f"🔐 Примеры используемых: {', '.join(sample_used)}"
                if len(server_data['used_configs']) > 2:
                    text += f" ... и еще {len(server_data['used_configs']) - 2}\n"
                else:
                    text += "\n"
                    
            text += "\n"
        
        # Добавляем общую статистику по конфигам
        text += f"📈 <b>Общая статистика:</b>\n"
        text += f"📁 Всего конфигов в системе: {len(all_configs)}\n"
        
        # Подсчет по серверам из базы данных
        total_used = sum(len(server['used_configs']) for server in servers_db.values())
        total_available = sum(len(server['available_configs']) for server in servers_db.values())
        text += f"🔐 Всего используется: {total_used}\n"
        text += f"📦 Всего доступно: {total_available}\n"
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в manage_servers: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        
@bot.message_handler(func=lambda message: message.text == '🔄 Синхронизировать конфиги' and is_admin(message.from_user.id))
def sync_configs(message):
    """Синхронизирует конфиги между файловой системой и базой данных"""
    try:
        # Получаем все конфиги из файловой системы
        all_configs = [f for f in os.listdir() if f.endswith('.conf')]
        
        for server_key, server_data in servers_db.items():
            server_name = server_data['name']
            
            # Определяем паттерн для поиска конфигов этого сервера
            if 'Germany' in server_name or 'DE' in server_name or '🇩🇪' in server_name:
                patterns = ['Grm', 'germany', 'Germany', 'DE']
            elif 'Netherlands' in server_name or 'NL' in server_name or '🇳🇱' in server_name:
                patterns = ['Ndr', 'netherlands', 'Netherlands', 'NL']
            else:
                patterns = [server_name.replace(' ', '_')]
            
            # Ищем конфиги по паттернам
            server_configs = []
            for config in all_configs:
                for pattern in patterns:
                    if pattern in config:
                        server_configs.append(config)
                        break
            
            # Обновляем доступные конфиги (те, что не используются)
            used_configs_list = list(server_data['used_configs'].values())
            available_configs = [cfg for cfg in server_configs if cfg not in used_configs_list]
            
            server_data['available_configs'] = available_configs
        
        save_data_to_file()
        
        bot.send_message(
            message.chat.id,
            f"✅ Конфиги синхронизированы!\n"
            f"📁 Всего конфигов в системе: {len(all_configs)}\n"
            f"🖥 Серверов обновлено: {len(servers_db)}",
            reply_markup=admin_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка в sync_configs: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка синхронизации: {str(e)}")
        
@bot.message_handler(func=lambda message: message.text == '🔄 Перезагрузить конфиги' and is_admin(message.from_user.id))
def reload_configs(message):
    """Принудительно перезагружает все конфиги из файловой системы"""
    try:
        # Получаем все конфиги
        all_configs = [f for f in os.listdir() if f.endswith('.conf')]
        
        # Сбрасываем конфиги для всех серверов
        for server_key, server_data in servers_db.items():
            server_name_clean = server_data['name'].replace(' ', '_')
            server_configs = [f for f in all_configs if server_name_clean in f]
            
            # Разделяем на используемые и доступные
            used_configs = {}
            available_configs = []
            
            for config in server_configs:
                # Проверяем, используется ли конфиг
                user_id = None
                for uid, data in users_db.items():
                    if 'subscriptions' in data:
                        for sub in data['subscriptions']:
                            if sub.get('config_file') == config:
                                user_id = uid
                                break
                    if user_id:
                        break
                
                if user_id:
                    used_configs[str(user_id)] = config
                else:
                    available_configs.append(config)
            
            server_data['available_configs'] = available_configs
            server_data['used_configs'] = used_configs
        
        save_data_to_file()
        
        bot.send_message(
            message.chat.id,
            f"✅ Конфиги перезагружены!\n"
            f"📁 Найдено конфигов: {len(all_configs)}\n"
            f"🖥 Серверов обновлено: {len(servers_db)}",
            reply_markup=admin_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка в reload_configs: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка перезагрузки конфигов: {str(e)}")

@bot.message_handler(func=lambda message: message.text == '🧾 Проверить платежи' and is_admin(message.from_user.id))
def check_payments(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        pending_payments = [pid for pid, p in payments_db.items() if p.get('status') == 'pending']
        
        if not pending_payments:
            bot.send_message(message.chat.id, "Нет платежей для проверки.", reply_markup=admin_menu_keyboard())
            return
        
        text = "📋 Платежи для проверки:\n\n"
        for pid in pending_payments[-5:]:  # Показываем последние 5 платежей
            p = payments_db[pid]
            text += f"#{pid}\nПользователь: @{p.get('username', 'N/A')}\nСумма: {p.get('amount', 'N/A')}\nБанк: {p.get('bank', 'N/A')}\n\n"
        
        # Добавляем кнопку "Назад" если больше нет платежей
        if len(pending_payments) <= 1:
            keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
            keyboard.add(
                types.KeyboardButton(f'✅ Подтвердить {pending_payments[-1]}'),
                types.KeyboardButton(f'❌ Отклонить {pending_payments[-1]}'),
                types.KeyboardButton('🔙 Назад')
            )
        else:
            keyboard = payment_verification_keyboard(pending_payments[-1])
        
        bot.send_message(message.chat.id, text, reply_markup=keyboard)
    except Exception as e:
        logger.error(f"Ошибка в check_payments: {e}")

@bot.message_handler(func=lambda message: message.text == '✅ Я оплатил')
def payment_done(message):
    try:
        user_id = message.from_user.id
        
        # Ищем последний ожидающий платеж пользователя
        user_payments = [p for pid, p in payments_db.items() 
                        if str(p.get('user_id')) == str(user_id) 
                        and p.get('status') == 'pending']
        
        if not user_payments:
            bot.send_message(user_id, "У вас нет ожидающих платежей.")
            return
        
        # Берем последний платеж
        payment = user_payments[-1]
        payment_id = next(pid for pid, p in payments_db.items() if p == payment)
        
        # Отправляем сообщение с инструкцией и ссылками
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('📲 Установить приложение'))
        markup.add(types.KeyboardButton('🔙 Назад'))
        
        try:
            with open('amnesia.png', 'rb') as photo:
                bot.send_photo(
                    user_id,
                    photo,
                    caption="""ПЛАТЕЖ НА ПРОВЕРКЕ АДМИНИСТРАТОРА
обычно это занимает пару минут
""",
                    reply_markup=markup
                )
        except FileNotFoundError:
            bot.send_message(
                user_id,
                """ПЛАТЕЖ НА ПРОВЕРКЕ АДМИНИСТРАТОРА
обычно это занимает пару минут
""",
                reply_markup=markup
            )
            
        # Уведомляем администратора
        bot.send_message(
            ADMIN_ID,
            f"🔔 Пользователь @{message.from_user.username} (ID: {user_id}) отправил платеж #{payment_id} на проверку"
        )
        
    except Exception as e:
        logger.error(f"Ошибка в payment_done: {e}")
        
# Модифицируем функцию approve_payment
@bot.message_handler(func=lambda message: message.text.startswith('✅ Подтвердить') and is_admin(message.from_user.id))
def approve_payment(message):
    try:
        payment_id = message.text.split()[-1]
        
        if payment_id not in payments_db:
            bot.send_message(message.chat.id, "Платеж не найден.")
            return
        
        payment = payments_db[payment_id]
        payment['status'] = 'approved'
        payment['approved_by'] = message.from_user.id
        payment['approved_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_payment(payment_id, payment)
        
        server_name = payment['server'].split(' (')[0]
        user_id = payment['user_id']
        user_data = get_user_data(user_id)
        
        # Проверяем, это продление или новая подписка
        is_extension = 'selected_config' in user_data
        
        if is_extension:
            # Продление существующей подписки
            config_file = user_data['selected_config']
            # Находим подписку для обновления
            for sub in user_data.get('subscriptions', []):
                if sub['config_file'] == config_file:
                    # Сбрасываем ВСЕ флаги уведомлений при продлении
                    if 'last_warnings' in sub:
                        del sub['last_warnings']
                    if 'expiry_notification_sent' in sub:
                        del sub['expiry_notification_sent']
                    
                    # Рассчитываем новую дату окончания
                    duration = payment['duration']
                    days = SUBSCRIPTION_PLANS.get(duration, {}).get('days', 30)
                    current_expiry = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if current_expiry > datetime.now():  # Если подписка еще активна, добавляем к текущей дате
                        new_expiry = current_expiry + timedelta(days=days)
                    else:  # Если подписка истекла, начинаем с текущей даты
                        new_expiry = datetime.now() + timedelta(days=days)
                    
                    sub['expiry_date'] = new_expiry.strftime("%Y-%m-%d %H:%M:%S")
                    sub['payment_id'] = payment_id
                    expiry_date = new_expiry.strftime("%Y-%m-%d %H:%M:%S")
                    break
        else:
            # Новая подписка
            config_file = get_random_config(server_name, user_id)
            if not config_file:
                bot.send_message(message.chat.id, f"Нет доступных конфигов для сервера {server_name}!")
                return
            
            # Рассчитываем дату окончания подписки
            duration = payment['duration']
            days = SUBSCRIPTION_PLANS.get(duration, {}).get('days', 30)
            expiry_date = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
            
            # Сохраняем информацию о подписке
            if 'subscriptions' not in user_data:
                user_data['subscriptions'] = []
                
            user_data['subscriptions'].append({
                'server': server_name,
                'config_file': config_file,
                'purchase_date': payment['approved_at'],
                'expiry_date': expiry_date,
                'payment_id': payment_id
            })
        
        user_data['username'] = payment.get('username')
        save_user_data(user_id, user_data)
        
        # 1. Отправляем пользователю информацию о подписке
        bot.send_message(
            user_id,
            f"""✅ Платеж #{payment_id} подтвержден!

📋 Информация о подписке:
🖥 Сервер: {server_name}
⏳ Срок: {payment['duration']}
📅 Действует до: {expiry_date if not is_extension else new_expiry.strftime("%Y-%m-%d %H:%M:%S")}

"""
        )
        
        # 2. Отправляем сам файл конфигурации с инструкцией (БЕЗ КНОПКИ)
        config_text = """1. Сохраните файл (ключ) на телефон

2. Откройте приложение AmneziaWG и нажмите "+"

3. Выберите "Создать из файла" и выберите сохраненный файл

4. Введите имя подключения "you name bot VPN" и "Сохранить"""
        
        with open(config_file, 'rb') as f:
            bot.send_document(
                user_id, 
                (os.path.basename(config_file), f),
                caption=config_text
            )
        
        # 3. Отправляем QR-код с инструкцией (БЕЗ КНОПКИ)
        # Читаем содержимое конфига для QR-кода
        with open(config_file, 'r') as f:
            config_content = f.read()
        
        # Генерируем QR-код
        qr_filename = f"{config_file}_qr.png"
        if generate_qr_code(config_content, qr_filename):
            # Отправляем QR-код с инструкцией
            qr_text = """ИЛИ

1. Отсканируйте QR-код (нажмите "Создать из QR-кода")

2. Введите имя подключения "you name bot VPN"

3. Нажмите "Сохранить"""
            
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption=qr_text
                )
        
        # 4. Отправляем финальное сообщение с инлайн кнопкой видеоинструкции
        final_markup = types.InlineKeyboardMarkup()
        final_markup.add(types.InlineKeyboardButton("📺 Смотреть видеоинструкцию", url="https://t.me/you name bot_09/38"))
        
        bot.send_message(
            user_id,
            "📹 Для подробной настройки посмотрите видеоинструкцию:",
            reply_markup=final_markup
        )
        
        # 5. Добавляем кнопку для установки приложения
        app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        app_markup.add(types.KeyboardButton('📲 Установить приложение'))
        app_markup.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(
            user_id,
            "Выберите следующее действие:",
            reply_markup=app_markup
        )
        
        bot.send_message(message.chat.id, f"Платеж #{payment_id} подтвержден. Пользователь получил конфиг: {config_file}")
    except Exception as e:
        logger.error(f"Ошибка в approve_payment: {e}")
        bot.send_message(
            ADMIN_ID,
            f"✅ Подписка продлена! Пользователь @{payment.get('username')} "
            f"получил +{days} дней. Конфиг: {config_file}"
        )
        
    except Exception as e:
        logger.error(f"Ошибка в approve_payment: {e}")

@bot.message_handler(func=lambda message: message.text.startswith('❌ Отклонить') and is_admin(message.from_user.id))
def reject_payment(message):
    try:
        payment_id = message.text.split()[-1]
        
        if payment_id not in payments_db:
            bot.send_message(message.chat.id, "Платеж не найден.")
            return
        
        payment = payments_db[payment_id]
        payment['status'] = 'rejected'
        payment['rejected_by'] = message.from_user.id
        payment['rejected_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_payment(payment_id, payment)
        
        try:
            bot.send_message(payment['user_id'], f"""
❌ Платеж #{payment_id} отклонен!

Проверьте чек и попробуйте снова.
""")
        except Exception as e:
            logger.error(f"Ошибка уведомления пользователя: {e}")
        
        bot.send_message(message.chat.id, f"Платеж #{payment_id} отклонен.")
    except Exception as e:
        logger.error(f"Ошибка в reject_payment: {e}")

@bot.message_handler(func=lambda message: message.text == '📢 Рассылка' and is_admin(message.from_user.id))
def broadcast_menu(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        msg = bot.send_message(message.chat.id, "Отправьте сообщение для рассылки (текст, фото, видео, документ):")
        bot.register_next_step_handler(msg, process_broadcast_message)
    except Exception as e:
        logger.error(f"Ошибка в broadcast_menu: {e}")

def process_broadcast_message(message):
    try:
        if message.text == '/cancel':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        users = list(users_db.keys())
        success = 0
        failed = 0
        
        for user_id in users:
            try:
                if message.content_type == 'text':
                    bot.send_message(user_id, message.text)
                elif message.content_type == 'photo':
                    bot.send_photo(user_id, message.photo[-1].file_id, caption=message.caption)
                elif message.content_type == 'video':
                    bot.send_video(user_id, message.video.file_id, caption=message.caption)
                elif message.content_type == 'document':
                    bot.send_document(user_id, message.document.file_id, caption=message.caption)
                elif message.content_type == 'audio':
                    bot.send_audio(user_id, message.audio.file_id, caption=message.caption)
                elif message.content_type == 'voice':
                    bot.send_voice(user_id, message.voice.file_id, caption=message.caption)
                
                success += 1
                time.sleep(0.1)
            except Exception as e:
                failed += 1
        
        bot.send_message(message.chat.id, f"""
📢 Результаты рассылки:
✅ Успешно: {success}
❌ Не удалось: {failed}
""", reply_markup=admin_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в process_broadcast_message: {e}")

@bot.message_handler(func=lambda message: message.text == '⚙️ Настройки оплаты' and is_admin(message.from_user.id))
def payment_settings(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        text = "⚙️ Реквизиты:\n\n"
        for method in payment_methods.values():
            text += f"<b>{method['bank']}</b>\nКарта: <code>{method['card_number']}</code>\n\n"
        
        text += "Используйте команды:\n/set_payment - добавить новый способ\n/delete_payment - удалить способ \nПример названия банка писать так Сбербанк"
        
        bot.send_message(message.chat.id, text, parse_mode='HTML')
    except Exception as e:
        logger.error(f"Ошибка в payment_settings: {e}")

@bot.message_handler(func=lambda message: message.text == '🗂 Управление конфигами' and is_admin(message.from_user.id))
def config_management(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        bot.send_message(message.chat.id, "Управление конфигурациями:", reply_markup=config_management_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в config_management: {e}")

# Добавляем команду для загрузки конфигов
@bot.message_handler(func=lambda message: message.text == '📤 Загрузить новый конфиг' and is_admin(message.from_user.id))
def upload_config(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        for server in servers_db.values():
            keyboard.add(types.KeyboardButton(f"📥 Загрузить для {server['name']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(message.chat.id, "Выберите сервер для загрузки конфига:", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_config_upload)
    except Exception as e:
        logger.error(f"Ошибка в upload_config: {e}")

def process_config_upload(message):
    try:
        if not message.text:
            bot.send_message(message.chat.id, "❌ Не получен текст сообщения.")
            return
            
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        server_name = message.text.replace('📥 Загрузить для ', '')
        msg = bot.send_message(message.chat.id, f"Отправьте файл конфигурации для сервера {server_name}:", reply_markup=types.ReplyKeyboardRemove())
        bot.register_next_step_handler(msg, lambda m: save_config_file(m, server_name))
    except Exception as e:
        logger.error(f"Ошибка в process_config_upload: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при загрузке конфига")

def save_config_file(message, server_name):
    try:
        if message.content_type != 'document':
            bot.send_message(message.chat.id, "Пожалуйста, отправьте файл конфигурации.")
            return
        
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        # Сохраняем файл с оригинальным именем
        original_filename = message.document.file_name
        config_filename = original_filename
        
        with open(config_filename, 'wb') as new_file:
            new_file.write(downloaded_file)
        
        # Добавляем конфиг в доступные для сервера
        server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
        servers_db[server_key]['available_configs'].append(config_filename)
        save_data_to_file()
        
        bot.send_message(message.chat.id, f"Конфигурация для сервера {server_name} успешно сохранена как {config_filename}!", reply_markup=admin_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в save_config_file: {e}")

@bot.message_handler(func=lambda message: message.text == '🗑 Удалить конфиг' and is_admin(message.from_user.id))
def delete_config(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        configs = [f for f in os.listdir() if f.endswith('.conf')]
        if not configs:
            bot.send_message(message.chat.id, "Нет конфигураций для удаления.")
            return
        
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        for config in configs:
            keyboard.add(types.KeyboardButton(config))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(message.chat.id, "Выберите конфиг для удаления:", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_config_deletion)
    except Exception as e:
        logger.error(f"Ошибка в delete_config: {e}")

def process_config_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        config_name = message.text
        if os.path.exists(config_name):
            os.remove(config_name)
            
            # Удаляем из базы серверов
            for key, server in list(servers_db.items()):
                if server['config'] == config_name:
                    del servers_db[key]
            
            save_data_to_file()
            bot.send_message(message.chat.id, f"Конфигурация {config_name} успешно удалена!", reply_markup=admin_menu_keyboard())
        else:
            bot.send_message(message.chat.id, "Файл не найден.")
    except Exception as e:
        logger.error(f"Ошибка в process_config_deletion: {e}")

@bot.message_handler(commands=['set_payment'])
def set_payment_method(message):
    if not is_admin(message.from_user.id):
        return
    
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        msg = bot.send_message(message.chat.id, """
Введите данные в формате:
Банк: Название
Карта: Номер
""")
        bot.register_next_step_handler(msg, process_payment_method)
    except Exception as e:
        logger.error(f"Ошибка в set_payment_method: {e}")

def process_payment_method(message):
    try:
        if message.text == '/cancel':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        lines = message.text.split('\n')
        method_data = {}
        
        for line in lines:
            if 'Банк:' in line:
                method_data['bank'] = line.split('Банк:')[1].strip()
            elif 'Карта:' in line:
                method_data['card_number'] = line.split('Карта:')[1].strip()
        
        method_key = method_data['bank'].lower().replace(' ', '_')
        payment_methods[method_key] = method_data
        save_data_to_file()
        
        # Рекомендация по названию файла изображения
        bank_name_lower = method_data['bank'].lower()
        image_recommendation = ""
        
        if 'сбер' in bank_name_lower:
            image_recommendation = "✅ Для этого банка будет использоваться изображение sber.png"
        elif 'альфа' in bank_name_lower:
            image_recommendation = "✅ Для этого банка будет использоваться изображение alfa.png"
        elif 'тинькофф' in bank_name_lower or 'тбанк' in bank_name_lower:
            image_recommendation = "✅ Для этого банка будет использоваться изображение tbank.PNG"
        else:
            image_recommendation = "ℹ️ Для этого банка будет использоваться стандартное изображение selectbank.png"
        
        bot.send_message(message.chat.id, 
                        f"✅ Реквизиты обновлены!\n{image_recommendation}", 
                        reply_markup=admin_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в process_payment_method: {e}")

@bot.message_handler(content_types=['photo'])
def handle_receipt(message):
    try:
        user_id = message.from_user.id
        
        # Ищем последний ожидающий платеж пользователя
        user_payments = [p for pid, p in payments_db.items() 
                        if str(p.get('user_id')) == str(user_id) 
                        and p.get('status') == 'pending']
        
        if not user_payments:
            bot.send_message(user_id, "У вас нет ожидающих платежей.")
            return
        
        # Берем последний платеж
        payment = user_payments[-1]
        payment_id = next(pid for pid, p in payments_db.items() if p == payment)
        
        bot.send_photo(ADMIN_ID, message.photo[-1].file_id, caption=f"""
📸 Чек #{payment_id}
Пользователь: @{message.from_user.username}
Сумма: {payment.get('amount', 'N/A')}
Банк: {payment.get('bank', 'N/A')}
""")
        
        bot.send_message(user_id, f"""
📨 Чек #{payment_id} получен!
Ожидайте подтверждения.
""")
    except Exception as e:
        logger.error(f"Ошибка в handle_receipt: {e}")

# Обработчики действий с конфигами
@bot.message_handler(func=lambda message: message.text == '📲 Установить приложение')
def install_wireguard(message):
    try:
        bot.send_message(message.chat.id, "Скачайте WireGuard с официального сайта: https://www.wireguard.com/install/")
    except Exception as e:
        logger.error(f"Ошибка в install_wireguard: {e}")

# Модифицируем функцию import_config
@bot.message_handler(func=lambda message: message.text.startswith('⚙️ Импортировать'))
def import_config(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        user_data = get_user_data(user_id)
        if 'subscriptions' not in user_data or not user_data['subscriptions']:
            bot.send_message(user_id, "У вас нет активных конфигураций.")
            return
        
        last_sub = user_data['subscriptions'][-1]
        config_file = last_sub['config_file']
        
        if not os.path.exists(config_file):
            bot.send_message(user_id, "Файл конфигурации не найден.")
            return
        
        # Проверяем не истекла ли подписка
        expiry_date = datetime.strptime(last_sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
        is_expired = expiry_date < datetime.now()
        
        if is_expired:
            bot.send_message(user_id, "⚠️ Срок действия вашей подписки истек! Для возобновления работы приобретите новый доступ.")
            return
        
        # Читаем содержимое конфига для QR-кода
        with open(config_file, 'r') as f:
            config_content = f.read()
        
        # Генерируем QR-код
        qr_filename = f"{config_file}_qr.png"
        if generate_qr_code(config_content, qr_filename):
            # Отправляем QR-код
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption="📲 Отсканируйте этот QR-код в приложении Amnezia для быстрой настройки"
                )
        
        # Отправляем инструкцию
        instructions = """📲 <b>Инструкция по установке в AmneziaWG:</b>

1. <b>Способ 1: Сканирование QR-кода</b>
   • Откройте Amnezia
   • Нажмите "+" (Добавить подключение)
   • Выберите "Сканировать QR-код"
   • Наведите камеру на QR-код выше
   • Нажмите "Сохранить"

2. <b>Способ 2: Импорт из файла</b>
   • Нажмите "+" (Добавить подключение)
   • Выберите "Импорт из файла"
   • Найдите и выберите отправленный вам файл конфигурации
   • Введите имя подключения (например: "Мой VPN")
   • Нажмите "Сохранить"

3. <b>Подключение:</b>
   • Выберите созданное подключение в списке
   • Нажмите "Подключиться"
   • Разрешите запрос на создание VPN-подключения

🔹 <b>Важно:</b>
• При первом подключении может потребоваться 1-2 минуты для установки соединения
• Не удаляйте файл конфигурации - он может понадобиться для повторного импорта
• Для автоматического подключения включите "Автоподключение" в настройках AmneziaWG

📹 <b>Видеоинструкция:</b> https://t.me/you name bot_09/38
"""
        
        # Отправляем сам файл конфигурации
        with open(config_file, 'rb') as f:
            bot.send_document(
                user_id,
                f,
                caption=instructions,
                parse_mode='HTML',
                visible_file_name=os.path.basename(config_file))
        
        # Дополнительные кнопки для удобства
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(
            types.KeyboardButton('📲 Установить приложение'),
            types.KeyboardButton('💾 Скачать конфиг'),
            types.KeyboardButton('🔙 Назад')
        )
        
        bot.send_message(
            user_id,
            "Выберите следующее действие:",
            reply_markup=markup
        )
            
    except Exception as e:
        logger.error(f"Ошибка в import_config: {e}")
        bot.send_message(user_id, "Произошла ошибка при подготовке инструкции. Пожалуйста, попробуйте позже.")

@bot.message_handler(commands=['delete_payment'])
def delete_payment_method(message):
    if not is_admin(message.from_user.id):
        return
    
    try:
        if not payment_methods:
            bot.send_message(message.chat.id, "Нет способов оплаты для удаления.")
            return
            
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        for method in payment_methods.values():
            keyboard.add(types.KeyboardButton(f"❌ Удалить {method['bank']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "Выберите способ оплаты для удаления:",
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_payment_deletion)
    except Exception as e:
        logger.error(f"Ошибка в delete_payment_method: {e}")

def process_payment_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        bank_name = message.text.replace('❌ Удалить ', '')
        method_key = next(key for key, val in payment_methods.items() if val['bank'] == bank_name)
        
        if method_key:
            del payment_methods[method_key]
            save_data_to_file()
            bot.send_message(
                message.chat.id,
                f"Способ оплаты {bank_name} успешно удален!",
                reply_markup=admin_menu_keyboard()
            )
        else:
            bot.send_message(message.chat.id, "Способ оплаты не найден.")
    except Exception as e:
        logger.error(f"Ошибка в process_payment_deletion: {e}")
        
@bot.message_handler(func=lambda message: message.text == '👥 Список покупателей' and is_admin(message.from_user.id))
def customers_list(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Создаем клавиатуру с новыми кнопками
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('🔍 Просроченные подписки'))
        keyboard.add(types.KeyboardButton('🗑 Удалить конфиг пользователя'))
        keyboard.add(types.KeyboardButton('🔄 Обновить список'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        # Формируем список активных пользователей (только НЕ просроченные)
        active_users = []
        current_time = datetime.now()
        
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                for sub in user_data['subscriptions']:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if expiry_date > current_time:  # Только активные подписки
                        days_left = (expiry_date - current_time).days
                        active_users.append({
                            'user_id': user_id,
                            'username': user_data.get('username', 'N/A'),
                            'server': sub['server'],
                            'config': sub['config_file'],
                            'expiry_date': sub['expiry_date'],
                            'days_left': days_left,
                            'purchase_date': sub.get('purchase_date', 'N/A')
                        })
        
        # Сортируем по количеству оставшихся дней
        active_users.sort(key=lambda x: x['days_left'])
        
        # Разбиваем сообщение на части если слишком длинное
        if not active_users:
            text = "📋 <b>Активные подписки:</b>\n\n"
            text += "Нет активных подписок."
            bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        else:
            # Отправляем общее количество
            bot.send_message(
                message.chat.id, 
                f"📋 <b>Активные подписки:</b>\n\nВсего активных подписок: {len(active_users)}",
                parse_mode='HTML',
                reply_markup=keyboard
            )
            
            # Разбиваем на части по 10 пользователей
            chunk_size = 10
            for i in range(0, len(active_users), chunk_size):
                chunk = active_users[i:i + chunk_size]
                text = f"📋 <b>Активные подписки (часть {i//chunk_size + 1}):</b>\n\n"
                
                for user in chunk:
                    text += (f"👤 <b>Пользователь:</b> @{user['username']} (ID: {user['user_id']})\n"
                            f"🖥 <b>Сервер:</b> {user['server']}\n"
                            f"🔑 <b>Конфиг:</b> {user['config']}\n"
                            f"📅 <b>Куплено:</b> {user['purchase_date']}\n"
                            f"⏳ <b>Осталось дней:</b> {user['days_left']}\n"
                            f"────────────────────\n")
                
                # Проверяем длину сообщения и отправляем
                if len(text) > 4000:
                    # Если все еще слишком длинное, разбиваем еще больше
                    lines = text.split('\n')
                    current_chunk = ""
                    for line in lines:
                        if len(current_chunk + line + '\n') > 4000:
                            bot.send_message(message.chat.id, current_chunk, parse_mode='HTML')
                            current_chunk = line + '\n'
                        else:
                            current_chunk += line + '\n'
                    if current_chunk:
                        bot.send_message(message.chat.id, current_chunk, parse_mode='HTML')
                else:
                    bot.send_message(message.chat.id, text, parse_mode='HTML')
        
    except Exception as e:
        logger.error(f"Ошибка в customers_list: {e}")
        bot.send_message(message.chat.id, f"Ошибка при получении списка: {str(e)}")
        
# Модифицируем функцию просроченных подписок
@bot.message_handler(func=lambda message: message.text == '🔍 Просроченные подписки' and is_admin(message.from_user.id))
def expired_subscriptions(message):
    try:
        expired_users = []
        current_time = datetime.now()
        
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                for sub in user_data['subscriptions']:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if expiry_date < current_time:
                        expired_days = (current_time - expiry_date).days
                        expired_users.append({
                            'user_id': user_id,
                            'username': user_data.get('username', 'N/A'),
                            'server': sub['server'],
                            'config': sub['config_file'],
                            'expiry_date': sub['expiry_date'],
                            'expired_days': expired_days,
                            'purchase_date': sub.get('purchase_date', 'N/A')
                        })
        
        # Сортируем по количеству дней просрочки
        expired_users.sort(key=lambda x: x['expired_days'], reverse=True)
        
        if not expired_users:
            bot.send_message(message.chat.id, "✅ Нет пользователей с просроченными подписками.")
            return
        
        # Создаем клавиатуру с кнопками для управления
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('🗑 Очистить просроченные'))
        keyboard.add(types.KeyboardButton('🔄 Обновить список'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        text = "⚠️ <b>Просроченные подписки:</b>\n\n"
        for user in expired_users:
            text += (f"👤 <b>Пользователь:</b> @{user['username']} (ID: {user['user_id']})\n"
                    f"🖥 <b>Сервер:</b> {user['server']}\n"
                    f"🔑 <b>Конфиг:</b> {user['config']}\n"
                    f"📅 <b>Куплено:</b> {user['purchase_date']}\n"
                    f"📅 <b>Истекла:</b> {user['expiry_date']} ({user['expired_days']} дней назад)\n"
                    f"────────────────────\n")
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка в expired_subscriptions: {e}")
        
@bot.message_handler(func=lambda message: message.text == '🔄 Обновить список' and is_admin(message.from_user.id))
def refresh_lists(message):
    """Обновляет списки подписок"""
    try:
        # Просто вызываем соответствующую функцию в зависимости от текущего контекста
        if 'просроченные' in message.text.lower() or '🔍' in message.text:
            expired_subscriptions(message)
        else:
            customers_list(message)
    except Exception as e:
        logger.error(f"Ошибка в refresh_lists: {e}")
        bot.send_message(message.chat.id, "Ошибка при обновлении списка.")
        
@bot.message_handler(func=lambda message: message.text == '🗑 Очистить просроченные' and is_admin(message.from_user.id))
def cleanup_expired(message):
    try:
        deleted_configs = []
        freed_servers = set()
        
        for user_id, user_data in list(users_db.items()):
            if 'subscriptions' in user_data:
                # Создаем новый список без просроченных подписок
                active_subs = []
                for sub in user_data['subscriptions']:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if expiry_date > datetime.now():
                        active_subs.append(sub)
                    else:
                        # Освобождаем конфиг на сервере
                        server_key = next(k for k, v in servers_db.items() if v['name'] == sub['server'])
                        if server_key and sub['config_file'] in servers_db[server_key]['used_configs'].values():
                            # Переносим конфиг обратно в доступные
                            servers_db[server_key]['available_configs'].append(sub['config_file'])
                            # Удаляем из используемых
                            user_id_key = str(user_id)
                            if user_id_key in servers_db[server_key]['used_configs']:
                                del servers_db[server_key]['used_configs'][user_id_key]
                            
                            deleted_configs.append(sub['config_file'])
                            freed_servers.add(sub['server'])
                
                # Обновляем подписки пользователя
                if active_subs:
                    users_db[user_id]['subscriptions'] = active_subs
                else:
                    del users_db[user_id]['subscriptions']
        
        save_data_to_file()
        
        text = "✅ <b>Очистка завершена:</b>\n\n"
        if deleted_configs:
            text += f"🗑 <b>Удалено конфигов:</b> {len(deleted_configs)}\n"
            text += f"🖥 <b>Затронуто серверов:</b> {len(freed_servers)}\n"
            text += "\n<b>Освобождённые конфиги:</b>\n"
            for config in deleted_configs:
                text += f"• {config}\n"
        else:
            text += "Не найдено просроченных подписок для очистки."
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка в cleanup_expired: {e}")
        bot.send_message(message.chat.id, f"Ошибка при очистке: {str(e)}", reply_markup=admin_menu_keyboard())
                
@bot.message_handler(func=lambda message: message.text == 'Активировать промокод')
def promo_code_handler(message):
    try:
        # Проверяем подписку
        if not check_subscription(message.from_user.id):
            bot.send_message(
                message.chat.id,
                "👋 Для использования бота необходимо подписаться на наши каналы:",
                reply_markup=subscription_keyboard()
            )
            return
            
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "✏️ Введите промокод:",
            reply_markup=markup
        )
        bot.register_next_step_handler(msg, process_promo_code)
    except Exception as e:
        logger.error(f"Ошибка в promo_code_handler: {e}")

def process_promo_code(message):
    try:
        user_id = message.from_user.id
        
        if message.text == '🔙 Назад':
            bot.send_message(user_id, "Возвращаемся в главное меню", reply_markup=main_menu_keyboard())
            return
        
        # Проверяем, что сообщение содержит текст, а не фото или другой контент
        if not message.text or not hasattr(message, 'text'):
            bot.send_message(user_id, "❌ Пожалуйста, введите текстовый промокод.")
            return
        
        if message.content_type == 'photo':
            # Обработка скриншота
            bot.send_photo(
                ADMIN_ID,
                message.photo[-1].file_id,
                caption=f"🔔 Запрос на активацию промокода от @{message.from_user.username} (ID: {user_id})"
            )
            
            # 1. Отправляем сообщение об установке приложения
            try:
                with open('amnesia.png', 'rb') as photo:
                    bot.send_photo(
                        user_id,
                        photo,
                        caption="Установите приложение AmneziaWG на свой смартфон"
                    )
            except FileNotFoundError:
                bot.send_message(
                    user_id,
                    "Установите приложение AmneziaWG на свой смартфон"
                )
            
            # 2. Отправляем ключ (смайлик)
            time.sleep(5)  # Задержка 5 секунд
            bot.send_message(user_id, "🔑")
            
            # 3. Создаем inline-клавиатуру с видеоинструкцией и скачиванием приложения
            inline_markup = types.InlineKeyboardMarkup()
            inline_markup.row(
                types.InlineKeyboardButton("📺 Смотреть видеоинструкцию", url="https://t.me/you name bot_09/38")
            )
            inline_markup.row(
                types.InlineKeyboardButton("🔗 ссылка на приложение", callback_data="download_app_menu")
            )
            
            bot.send_message(
                user_id,
                "✅ Ваш запрос на активацию промокода получен и отправлен администратору.\n\n"
                "Ожидайте подтверждения. Обычно это занимает несколько минут.",
                reply_markup=inline_markup
            )
            
            # 4. Добавляем обычную клавиатуру с кнопкой назад
            reply_keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            reply_keyboard.add(types.KeyboardButton('🔙 Назад'))
            
            bot.send_message(
                user_id,
                "Выберите следующее действие:",
                reply_markup=reply_keyboard
            )
            return
        
        promo_code = message.text.strip().upper()
        
        # Проверяем, что промокод не пустой
        if not promo_code:
            bot.send_message(user_id, "❌ Промокод не может быть пустым!")
            return
            
        # Проверяем, активировал ли пользователь уже этот промокод
        user_data = get_user_data(user_id)
        if 'used_promo_codes' in user_data and promo_code in user_data['used_promo_codes']:
            bot.send_message(user_id, "❌ Вы уже активировали этот промокод ранее!")
            return
            
        if promo_code not in PROMO_CODES:
            bot.send_message(user_id, "⚠️ вводить только заглавными буквами")
            return
            
        promo_data = PROMO_CODES[promo_code]
        server_name = promo_data['server']
        
        # Получаем случайный конфиг для выбранного сервера
        config_file = get_random_config(server_name, user_id)
        if not config_file:
            bot.send_message(user_id, "⚠️ На сервере закончились свободные конфигурации. Попробуйте позже.")
            return
            
        expiry_date = (datetime.now() + timedelta(days=promo_data['days'])).strftime("%Y-%m-%d %H:%M:%S")
        
        # Сохраняем информацию о подписке
        if 'subscriptions' not in user_data:
            user_data['subscriptions'] = []
            
        user_data['subscriptions'].append({
            'server': server_name,
            'config_file': config_file,
            'purchase_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'expiry_date': expiry_date,
            'type': 'promo',
            'promo_code': promo_code
        })
        
        # Добавляем промокод в список использованных
        if 'used_promo_codes' not in user_data:
            user_data['used_promo_codes'] = []
        user_data['used_promo_codes'].append(promo_code)
        
        save_user_data(user_id, user_data)
        
        # 1. Сначала отправляем сообщение об установке приложения с фото
        try:
            with open('amnesia.png', 'rb') as photo:
                bot.send_photo(
                    user_id,
                    photo,
                    caption="Установите приложение AmneziaWG на свой смартфон"
                )
        except FileNotFoundError:
            bot.send_message(
                user_id,
                "Установите приложение AmneziaWG на свой смартфон"
            )
        
        # 2. Отправляем ключ (смайлик) с задержкой 5 секунд
        time.sleep(5)
        bot.send_message(user_id, "🔑")
        
        # 3. Отправляем сообщение об успешной активации промокода
        server_info_text = f"""🎉 Промокод активирован!

📋 Информация о подписке:
🌍 Сервер: {server_name}
⏳ Срок: {promo_data['days']} дней
📅 Активен до: {expiry_date}"""

        bot.send_message(user_id, server_info_text)
        
        # 4. Отправляем конфиг пользователю с инструкцией
        config_text = """1. Сохраните файл (ключ) на телефон

2. Откройте приложение AmneziaWG и нажмите "+"

3. Выберите "Создать из файла" и выберите сохраненный файл

4. Введите имя подключения "you name bot VPN" и "Сохранить" """
        
        with open(config_file, 'rb') as f:
            bot.send_document(
                user_id,
                f,
                caption=config_text,
                visible_file_name=os.path.basename(config_file)
            )
        
        # 5. Генерируем и отправляем QR-код
        with open(config_file, 'r') as f:
            config_content = f.read()
        
        qr_filename = f"{config_file}_qr.png"
        if generate_qr_code(config_content, qr_filename):
            qr_text = """ИЛИ

1. Отсканируйте QR-код (нажмите "Создать из QR-кода")
2. Введите имя подключения "you name bot VPN"
3. Нажмите "Сохранить" """
            
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption=qr_text
                )
        
        # 6. Отправляем inline-клавиатуру с видеоинструкцией и скачиванием приложения
        time.sleep(2)
        bot.send_message(user_id, "🧰")
        
        inline_markup = types.InlineKeyboardMarkup()
        inline_markup.row(
            types.InlineKeyboardButton("📺 Смотреть видеоинструкцию", url="https://t.me/you name bot_09/38")
        )
        inline_markup.row(
            types.InlineKeyboardButton("🔗 ссылка на приложение", callback_data="download_app_menu")
        )
        
        final_text = """ЕСЛИ НЕ СМОГЛИ АКТИВИРОВАТЬ:"""
        
        bot.send_message(
            user_id,
            final_text,
            reply_markup=inline_markup
        )
        
        # 7. Добавляем обычную клавиатуру с кнопкой назад (остаемся в этом меню)
        reply_keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(types.KeyboardButton('🔙 Назад'))
        
            
        # Уведомляем админа
        bot.send_message(
            ADMIN_ID,
            f"🔔 Промокод {promo_code} активирован пользователем @{message.from_user.username} (ID: {user_id})"
        )
        
    except Exception as e:
        logger.error(f"Ошибка в process_promo_code: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        bot.send_message(message.chat.id, "❌ Произошла ошибка при обработке промокода. Пожалуйста, попробуйте позже.")
        
        
@bot.callback_query_handler(func=lambda call: call.data.startswith('extend:'))
def handle_extend_callback(call):
    try:
        user_id = call.from_user.id
        config_file = call.data.split(':')[1]
        
        # Находим подписку пользователя по файлу конфига
        user_data = get_user_data(user_id)
        subscription = None
        
        for sub in user_data.get('subscriptions', []):
            if sub['config_file'] == config_file:
                subscription = sub
                break
        
        if not subscription:
            bot.answer_callback_query(call.id, "Подписка не найдена!")
            return
            
        # Сохраняем выбранный сервер для продления
        user_data['selected_server'] = subscription['server']
        save_user_data(user_id, user_data)
        
        # Предлагаем выбрать срок продления
        bot.send_message(
            user_id,
            f"Вы выбрали продление подписки для сервера {subscription['server']}.\nВыберите срок продления:",
            reply_markup=duration_menu_keyboard()
        )
        
        bot.answer_callback_query(call.id)
    except Exception as e:
        logger.error(f"Ошибка в handle_extend_callback: {e}")
        bot.answer_callback_query(call.id, "Произошла ошибка!")
        
@bot.message_handler(func=lambda message: message.text == '🔑 Мои ключи')
def my_keys_handler(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        user_data = get_user_data(user_id)
        if 'subscriptions' not in user_data or not user_data['subscriptions']:
            bot.send_message(user_id, "У вас пока нет активных ключей.")
            return
        
        text = "🔑 Ваши VPN ключи:\n\n"
        for idx, sub in enumerate(reversed(user_data['subscriptions']), 1):
            expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
            is_expired = expiry_date < datetime.now()
            days_left = (expiry_date - datetime.now()).days if not is_expired else 0
            
            text += f"🔹 Ключ #{idx}\n"
            text += f"🖥 Сервер: {sub['server']}\n"
            text += f"📅 Создан: {sub['purchase_date']}\n"
            text += f"⏳ Действует до: {sub['expiry_date']}\n"
            
            if is_expired:
                text += "⚠️ Статус: Истек\n\n"
            else:
                text += f"✅ Статус: Активен ({days_left} дней осталось)\n\n"
        
        bot.send_message(user_id, text)
        
        # Отправляем файлы конфигураций и QR-коды
        for sub in user_data['subscriptions']:
            if os.path.exists(sub['config_file']):
                # Читаем содержимое конфига для QR-кода
                with open(sub['config_file'], 'r') as f:
                    config_content = f.read()
                
                # Генерируем QR-код
                qr_filename = f"{sub['config_file']}_qr.png"
                if generate_qr_code(config_content, qr_filename):
                    # Отправляем QR-код
                    with open(qr_filename, 'rb') as qr_file:
                        bot.send_photo(
                            user_id,
                            qr_file,
                            caption=f"QR-код для {sub['server']}"
                        )
                
                # Отправляем сам конфиг
                with open(sub['config_file'], 'rb') as f:
                    bot.send_document(
                        user_id,
                        f,
                        caption=f"Конфигурация для {sub['server']}",
                        visible_file_name=os.path.basename(sub['config_file'])
                    )
            else:
                bot.send_message(user_id, f"Файл конфигурации {sub['config_file']} не найден.")
                
    except Exception as e:
        logger.error(f"Ошибка в my_keys_handler: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка при получении списка ключей.")
        

# Модифицируем функцию download_config
@bot.message_handler(func=lambda message: message.text == '💾 Скачать конфиг')
def download_config(message):
    try:
        user_id = message.from_user.id
        user_data = get_user_data(user_id)
        
        if 'subscriptions' not in user_data or not user_data['subscriptions']:
            bot.send_message(user_id, "У вас нет активных подписок.")
            return
        
        last_sub = user_data['subscriptions'][-1]
        config_file = last_sub['config_file']
        
        if not os.path.exists(config_file):
            bot.send_message(user_id, "Файл конфигурации не найден.")
            return
        
        # Проверяем не истекла ли подписка
        expiry_date = datetime.strptime(last_sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
        is_expired = expiry_date < datetime.now()
        
        # Читаем содержимое конфига для QR-кода
        with open(config_file, 'r') as f:
            config_content = f.read()
        
        # Генерируем QR-код
        qr_filename = f"{config_file}_qr.png"
        if generate_qr_code(config_content, qr_filename):
            # Отправляем QR-код
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption="📲 Отсканируйте этот QR-код в приложении Amnezia для быстрой настройки"
                )
        
        caption = f"Ваш конфигурационный файл для {last_sub['server']}"
        if is_expired:
            caption += "\n⚠️ Срок действия истек! Для продления приобретите новый доступ."
        
        with open(config_file, 'rb') as f:
            bot.send_document(
                user_id,
                f,
                caption=caption,
                visible_file_name=os.path.basename(config_file))
            
    except Exception as e:
        logger.error(f"Ошибка в download_config: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка при отправке файла")

# Запуск бота
def run_bot():
    # Запускаем мониторинг подписок в отдельном потоке
    monitor_thread = threading.Thread(target=subscription_monitor, daemon=True)
    monitor_thread.start()
    
    logger.info("Бот запущен и готов к работе")
    
    while True:
        try:
            logger.info("Запуск polling...")
            # УБИРАЕМ restart_on_change=True или устанавливаем в False
            bot.infinity_polling(timeout=60, long_polling_timeout=60, restart_on_change=False)
        except requests.exceptions.ConnectionError:
            logger.error("Ошибка соединения. Повторная попытка через 15 секунд...")
            time.sleep(15)
        except requests.exceptions.ReadTimeout:
            logger.error("Таймаут соединения. Повторная попытка через 10 секунд...")
            time.sleep(10)
        except Exception as e:
            logger.error(f"Критическая ошибка бота: {e}")
            logger.error("Перезапуск через 30 секунд...")
            time.sleep(30)

@bot.message_handler(func=lambda message: message.text == '🗑 Удалить пользователя' and is_admin(message.from_user.id))
def delete_user_by_id(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        msg = bot.send_message(
            message.chat.id,
            "Введите ID пользователя для удаления:",
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, process_user_deletion)
    except Exception as e:
        logger.error(f"Ошибка в delete_user_by_id: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при удалении пользователя")

def process_user_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        user_id_to_delete = message.text.strip()
        
        # Проверяем, что введен корректный ID
        if not user_id_to_delete.isdigit():
            bot.send_message(message.chat.id, "❌ ID пользователя должен содержать только цифры!")
            return
        
        if user_id_to_delete not in users_db:
            bot.send_message(message.chat.id, "❌ Пользователь с таким ID не найден!")
            return
        
        user_data = users_db[user_id_to_delete]
        username = user_data.get('username', 'N/A')
        
        # Освобождаем все конфиги пользователя
        if 'subscriptions' in user_data:
            for sub in user_data['subscriptions']:
                server_name = sub['server']
                config_file = sub['config_file']
                
                # Находим сервер и освобождаем конфиг
                for server_key, server_data in servers_db.items():
                    if server_data['name'] == server_name:
                        # Переносим конфиг обратно в доступные
                        if config_file not in server_data['available_configs']:
                            server_data['available_configs'].append(config_file)
                        
                        # Удаляем из используемых
                        if user_id_to_delete in server_data['used_configs']:
                            del server_data['used_configs'][user_id_to_delete]
                        break
        
        # Удаляем пользователя из базы
        del users_db[user_id_to_delete]
        save_data_to_file()
        
        # Удаляем связанные платежи
        payments_to_delete = []
        for payment_id, payment_data in payments_db.items():
            if str(payment_data.get('user_id')) == user_id_to_delete:
                payments_to_delete.append(payment_id)
        
        for payment_id in payments_to_delete:
            del payments_db[payment_id]
        
        save_data_to_file()
        
        bot.send_message(
            message.chat.id,
            f"✅ Пользователь @{username} (ID: {user_id_to_delete}) успешно удален!\n"
            f"🗑 Освобождены все конфиги пользователя.",
            reply_markup=admin_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка в process_user_deletion: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при удалении пользователя", reply_markup=admin_menu_keyboard())

@bot.message_handler(func=lambda message: message.text == '🗑 Удалить конфиг пользователя' and is_admin(message.from_user.id))
def delete_user_config(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Получаем список пользователей с активными подписками
        active_users = []
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                for sub in user_data['subscriptions']:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if expiry_date > datetime.now():
                        active_users.append({
                            'user_id': user_id,
                            'username': user_data.get('username', 'N/A'),
                            'server': sub['server'],
                            'config': sub['config_file']
                        })
        
        if not active_users:
            bot.send_message(message.chat.id, "Нет активных пользователей для удаления.")
            return
        
        # Создаем клавиатуру с пользователями
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        for user in active_users:
            btn_text = f"👤 {user['username']} ({user['server']})"
            keyboard.add(types.KeyboardButton(btn_text))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "Выберите пользователя для удаления конфига:",
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_user_config_deletion)
    except Exception as e:
        logger.error(f"Ошибка в delete_user_config: {e}")

def process_user_config_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        # Парсим выбор пользователя
        username = message.text.split(' (')[0].replace('👤 ', '')
        server = message.text.split(' (')[1].replace(')', '')
        
        # Находим пользователя
        for user_id, user_data in users_db.items():
            if user_data.get('username') == username and 'subscriptions' in user_data:
                for sub in user_data['subscriptions']:
                    if sub['server'] == server:
                        # Освобождаем конфиг на сервере
                        server_key = next(k for k, v in servers_db.items() if v['name'] == sub['server'])
                        if server_key and sub['config_file'] in servers_db[server_key]['used_configs'].values():
                            servers_db[server_key]['available_configs'].append(sub['config_file'])
                            del servers_db[server_key]['used_configs'][str(user_id)]
                        
                        # Удаляем подписку
                        user_data['subscriptions'].remove(sub)
                        save_data_to_file()
                        
                        # Уведомляем пользователя
                        try:
                            bot.send_message(
                                user_id,
                                f"⚠️ Ваш VPN конфиг для сервера {sub['server']} был удален администратором. "
                                "Для возобновления работы необходимо приобрести новый доступ."
                            )
                        except Exception as e:
                            logger.error(f"Не удалось уведомить пользователя: {e}")
                        
                        bot.send_message(
                            message.chat.id,
                            f"Конфиг пользователя @{username} для сервера {sub['server']} успешно удален!",
                            reply_markup=admin_menu_keyboard()
                        )
                        return
        
        bot.send_message(message.chat.id, "Пользователь не найден.", reply_markup=admin_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в process_user_config_deletion: {e}")
        
@bot.callback_query_handler(func=lambda call: call.data == "download_app_menu")
def handle_download_app_callback(call):
    """Обработчик нажатия на кнопку скачивания приложения"""
    try:
        user_id = call.from_user.id
        
        # Создаем inline-клавиатуру со ссылками для скачивания
        markup = types.InlineKeyboardMarkup()
        markup.row(
            types.InlineKeyboardButton("Android", url="https://play.google.com/store/apps/details?id=org.amnezia.awg"),
            types.InlineKeyboardButton("iOS", url="https://apps.apple.com/ru/app/amneziawg/id6478942365")
        )
        markup.row(
            types.InlineKeyboardButton("Windows", url="https://github.com/amnezia-vpn/amneziawg-windows-client/releases/tag/1.0.2"),
            types.InlineKeyboardButton("macOS", url="https://apps.apple.com/us/app/amneziawg/id6478942365")
        )
        
        bot.send_message(
            user_id,
            "📥 Выберите вашу платформу для скачивания AmneziaWG:",
            reply_markup=markup
        )
        
        bot.answer_callback_query(call.id)
        
    except Exception as e:
        logger.error(f"Ошибка в handle_download_app_callback: {e}")
        bot.answer_callback_query(call.id, "❌ Ошибка при открытии меню скачивания")
# Добавляем команду для отключения конфига
@bot.message_handler(func=lambda message: message.text.startswith('🚫 Отключить конфиг') and is_admin(message.from_user.id))
def disable_config(message):
    try:
        config_file = message.text.replace('🚫 Отключить конфиг ', '')
        
        # Находим сервер и пользователя для этого конфига
        for server in servers_db.values():
            if config_file in server['used_configs'].values():
                user_id = next(uid for uid, cfg in server['used_configs'].items() if cfg == config_file)
                
                # Возвращаем конфиг в доступные
                server['available_configs'].append(config_file)
                del server['used_configs'][user_id]
                save_data_to_file()
                
                # Уведомляем пользователя
                try:
                    bot.send_message(user_id, f"""
⚠️ Ваш VPN конфиг {config_file} был отключен.
Для возобновления работы необходимо приобрести новый доступ.
""")
                except Exception as e:
                    logger.error(f"Не удалось уведомить пользователя {user_id}: {e}")
                
                bot.send_message(message.chat.id, f"Конфиг {config_file} успешно отключен.", reply_markup=admin_menu_keyboard())
                return
        
        bot.send_message(message.chat.id, "Конфиг не найден среди используемых.")
    except Exception as e:
        logger.error(f"Ошибка в disable_config: {e}")

def check_payment_images():
    """Проверяет существование файлов изображений для платежных систем"""
    required_images = ['sber.png', 'alfa.png', 'tbank.PNG', 'selectbank.png']
    missing_images = []
    
    for image in required_images:
        if not os.path.exists(image):
            missing_images.append(image)
    
    if missing_images:
        logger.warning(f"Отсутствуют изображения для платежных систем: {', '.join(missing_images)}")
        return False
    else:
        logger.info("Все изображения для платежных систем найдены")
        return True
# В функции process_config_deletion добавляем проверку на используемые конфиги
def process_config_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        config_name = message.text
        if os.path.exists(config_name):
            # Проверяем, не используется ли конфиг
            for server in servers_db.values():
                if config_name in server['used_configs'].values():
                    bot.send_message(message.chat.id, "Этот конфиг используется и не может быть удален!")
                    return
                
                if config_name in server['available_configs']:
                    server['available_configs'].remove(config_name)
            
            os.remove(config_name)
            save_data_to_file()
            bot.send_message(message.chat.id, f"Конфигурация {config_name} успешно удалена!", reply_markup=admin_menu_keyboard())
        else:
            bot.send_message(message.chat.id, "Файл не найден.")
    except Exception as e:
        logger.error(f"Ошибка в process_config_deletion: {e}")

if __name__ == '__main__':
    # Синхронизируем конфиги при запуске
    all_configs = [f for f in os.listdir() if f.endswith('.conf')]
    
    for server_key, server_data in servers_db.items():
        server_name = server_data['name']
        
        # Определяем паттерн для поиска конфигов этого сервера
        if 'Germany' in server_name or 'DE' in server_name or '🇩🇪' in server_name:
            patterns = ['Grm', 'germany', 'Germany', 'DE']
        elif 'Netherlands' in server_name or 'NL' in server_name or '🇳🇱' in server_name:
            patterns = ['Ndr', 'netherlands', 'Netherlands', 'NL']
        else:
            patterns = [server_name.replace(' ', '_')]
        
        # Ищем конфиги по паттернам
        server_configs = []
        for config in all_configs:
            for pattern in patterns:
                if pattern in config:
                    server_configs.append(config)
                    break
        
        # Обновляем доступные конфиги (те, что не используются)
        used_configs_list = list(server_data['used_configs'].values())
        available_configs = [cfg for cfg in server_configs if cfg not in used_configs_list]
        
        server_data['available_configs'] = available_configs
        
        logger.info(f"Сервер {server_name}: {len(available_configs)} доступных, {len(server_data['used_configs'])} используемых конфигов")
    
    save_data_to_file()
    run_bot()